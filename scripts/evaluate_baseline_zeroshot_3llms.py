"""
Évaluation ZERO-SHOT (baseline) — 3 LLMs sans fine-tuning
==========================================================
Modèles évalués séquentiellement sur les 450 exemples du test set :
  1. ALLaM-7B-Instruct      (humain-ai/ALLaM-7B-Instruct-preview)
  2. LLaMA 3-8B-Instruct    (meta-llama/Meta-Llama-3-8B-Instruct)
  3. Mistral-7B-Instruct    (mistralai/Mistral-7B-Instruct-v0.3)

Objectif : établir la baseline de référence AVANT tout fine-tuning.
Les résultats permettront de mesurer le gain apporté par chaque
configuration QLoRA testée (v3, v4, v4-r32, etc.).

Chaque modèle est chargé en QLoRA 4-bit NF4 (même quantification que
le fine-tuning) pour une comparaison équitable des conditions GPU.
Aucun adapteur LoRA n'est chargé.

Sorties par modèle :
  - baseline_zeroshot_<modele>.json   : métriques + prédictions
  - baseline_zeroshot_<modele>.xlsx   : rapport Excel multi-onglets

Sortie globale :
  - baseline_zeroshot_comparaison.json : tableau comparatif des 3 modèles

pip install transformers datasets bitsandbytes accelerate torch
            sacrebleu rouge-score openpyxl packaging
"""

import os, gc, json, random, datetime
import torch
import numpy as np

from transformers import AutoModelForCausalLM, AutoTokenizer, BitsAndBytesConfig

from sacrebleu.metrics import BLEU, TER, CHRF
from rouge_score import rouge_scorer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation

# ══════════════════════════════════════════════════════════════
# 0. CONFIGURATION GLOBALE
# ══════════════════════════════════════════════════════════════
os.environ["CUDA_VISIBLE_DEVICES"] = "0,1"
os.environ["TOKENIZERS_PARALLELISM"] = "false"
os.environ["HF_HUB_DISABLE_MODEL_CARD"] = "1"

# Dossier qui contient test_set_FINAL.json
# (produit par n'importe lequel des scripts de fine-tuning)
TEST_FILE  = r"C:\Users\infocom\tradiction_dialect\llama3_8b_tun_msa_output-3000\test_set_FINAL.json"
OUTPUT_DIR = r"C:\Users\infocom\tradiction_dialect\baseline_zeroshot"
SEED       = 42
MAX_NEW_TOKENS = 128

random.seed(SEED)
os.makedirs(OUTPUT_DIR, exist_ok=True)

SEP = "=" * 64

gc.collect()
torch.cuda.empty_cache()

print(SEP)
print("  ÉVALUATION ZERO-SHOT BASELINE — 3 LLMs — Tunisien → MSA")
print(SEP)
print(f"  PyTorch       : {torch.__version__}")
print(f"  CUDA dispo    : {torch.cuda.is_available()}")
print(f"  Nombre GPUs   : {torch.cuda.device_count()}")
if torch.cuda.is_available():
    for i in range(torch.cuda.device_count()):
        name = torch.cuda.get_device_name(i)
        vram = torch.cuda.get_device_properties(i).total_memory / 1e9
        print(f"  GPU {i}         : {name}  ({vram:.1f} Go VRAM)")
print(f"  Output dir    : {OUTPUT_DIR}")
print(SEP + "\n")

# ══════════════════════════════════════════════════════════════
# 1. CHARGEMENT DU TEST SET
# ══════════════════════════════════════════════════════════════
if not os.path.exists(TEST_FILE):
    raise FileNotFoundError(
        f"❌ Test set introuvable : {TEST_FILE}\n"
        "   Pointez TEST_FILE vers le test_set_FINAL.json produit\n"
        "   par l'un des scripts de fine-tuning."
    )

with open(TEST_FILE, "r", encoding="utf-8") as f:
    test_data = json.load(f)

print(f"✓ Test set chargé : {len(test_data)} exemples\n")

# ══════════════════════════════════════════════════════════════
# 2. DÉFINITION DES MODÈLES ET LEURS TEMPLATES ZERO-SHOT
# ══════════════════════════════════════════════════════════════
# Chaque modèle a son propre format de prompt natif.
# En zero-shot, on utilise directement les balises du modèle
# sans injecter d'exemples du dataset de fine-tuning.

MODELS = [
    {
        "name"         : "ALLaM-7B",
        "model_id"     : "humain-ai/ALLaM-7B-Instruct-preview",
        "short_name"   : "allam7b",
        "trust_remote" : True,
        # Format Alpaca-style (format natif ALLaM)
        "prompt_fn"    : lambda item: (
            "### Instruction:\n"
            f"{item['instruction'].strip()}\n\n"
            "### Input:\n"
            f"{item['input'].strip()}\n\n"
            "### Response:\n"
        ),
    },
    {
        "name"         : "LLaMA 3-8B",
        "model_id"     : "meta-llama/Meta-Llama-3-8B-Instruct",
        "short_name"   : "llama3_8b",
        "trust_remote" : False,
        # Format chat LLaMA 3 avec balises spéciales
        "prompt_fn"    : lambda item: (
            "<|begin_of_text|>"
            "<|start_header_id|>system<|end_header_id|>\n\n"
            f"{item['instruction'].strip()}<|eot_id|>"
            "<|start_header_id|>user<|end_header_id|>\n\n"
            f"{item['input'].strip()}<|eot_id|>"
            "<|start_header_id|>assistant<|end_header_id|>\n\n"
        ),
    },
    {
        "name"         : "Mistral-7B",
        "model_id"     : "mistralai/Mistral-7B-Instruct-v0.3",
        "short_name"   : "mistral7b",
        "trust_remote" : False,
        # Format [INST] natif Mistral
        "prompt_fn"    : lambda item: (
            f"<s>[INST] {item['instruction'].strip()}\n\n"
            f"{item['input'].strip()} [/INST]"
        ),
    },
]

# ══════════════════════════════════════════════════════════════
# 3. MÉTRIQUES — FONCTIONS COMMUNES
# ══════════════════════════════════════════════════════════════
_bleu_sent  = BLEU(effective_order=True, tokenize="char")
_bleu_corp  = BLEU(tokenize="char")
_ter_metric = TER()
_chrf_metric= CHRF(word_order=2)
_rouge      = rouge_scorer.RougeScorer(
    ["rouge1", "rouge2", "rougeL"], use_stemmer=False
)

def is_exact_match(pred, ref):
    return pred.strip().rstrip("،.") == ref.strip().rstrip("،.")

def sentence_bleu(pred, ref):
    try:    return _bleu_sent.sentence_score(pred, [ref]).score
    except: return 0.0

def sentence_ter(pred, ref):
    try:    return _ter_metric.sentence_score(pred, [ref]).score
    except: return 100.0

def sentence_chrf(pred, ref):
    try:    return _chrf_metric.sentence_score(pred, [ref]).score
    except: return 0.0

def sentence_rouge(pred, ref):
    try:
        s = _rouge.score(ref, pred)
        return {"rouge1": s["rouge1"].fmeasure,
                "rouge2": s["rouge2"].fmeasure,
                "rougeL": s["rougeL"].fmeasure}
    except:
        return {"rouge1": 0.0, "rouge2": 0.0, "rougeL": 0.0}

def token_f1(pred, ref):
    pt = set(pred.strip().split())
    rt = set(ref.strip().split())
    if not pt or not rt:
        return float(pt == rt)
    common = pt & rt
    if not common:
        return 0.0
    p = len(common) / len(pt)
    r = len(common) / len(rt)
    return 2 * p * r / (p + r)

# ══════════════════════════════════════════════════════════════
# 4. EXPORT EXCEL — FONCTION COMMUNE
# ══════════════════════════════════════════════════════════════
def _header_style(cell, bg="1F4E79"):
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFFFFF")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def _data_style(cell, wrap=False):
    cell.font      = Font(name="Arial", size=9)
    cell.alignment = Alignment(vertical="center", wrap_text=wrap, horizontal="right")
    thin = Side(style="thin", color="D9D9D9")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def _pct_bar(ws, col_letter, row_start, row_end):
    rule = DataBarRule(start_type="num", start_value=0,
                       end_type="num",   end_value=1, color="4472C4")
    ws.conditional_formatting.add(
        f"{col_letter}{row_start}:{col_letter}{row_end}", rule
    )

def export_excel(results, summary, model_display_name, out_path):
    wb = Workbook()

    # ── Onglet 1 : Métriques globales ────────────────────────
    ws_sum = wb.active
    ws_sum.title = "📋 Métriques Globales"

    ws_sum.merge_cells("A1:C1")
    t = ws_sum["A1"]
    t.value = f"{model_display_name} — Évaluation ZERO-SHOT — Tunisien → MSA"
    t.font  = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    t.fill  = PatternFill("solid", start_color="1F4E79")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 28

    kv = [
        ("Modèle",                  summary["model"]),
        ("Mode",                    "ZERO-SHOT (sans fine-tuning)"),
        ("Tâche",                   summary["task"]),
        ("Nombre exemples test",    summary["n_test"]),
        ("Date évaluation",         datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Exact Match Accuracy",    f"{summary['exact_match_acc']*100:.2f} %"),
        ("Token F1 (avg)",          f"{summary['avg_token_f1']:.4f}"),
        ("BLEU (corpus)",           f"{summary['corpus_bleu']:.2f}"),
        ("chrF++ (corpus)",         f"{summary['corpus_chrf_pp']:.2f}"),
        ("TER (corpus, ↓)",         f"{summary['corpus_ter']:.2f}"),
        ("ROUGE-1 (avg)",           f"{summary['avg_rouge1']:.4f}"),
        ("ROUGE-2 (avg)",           f"{summary['avg_rouge2']:.4f}"),
        ("ROUGE-L (avg)",           f"{summary['avg_rougeL']:.4f}"),
    ]
    for ri, (k, v) in enumerate(kv, 2):
        ck = ws_sum.cell(ri, 1, k)
        cv = ws_sum.cell(ri, 2, v)
        if k:
            ck.font = Font(bold=True, name="Arial", size=10)
            cv.font = Font(name="Arial", size=10)
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 50

    # ── Onglet 2 : Évaluation expert ─────────────────────────
    ws_exp = wb.create_sheet("🔍 Évaluation Expert")
    ws_exp.freeze_panes = "A2"

    hdrs = ["ID", "Texte Tunisien (input)", "Référence MSA",
            f"Zéro-shot {model_display_name}",
            "Exact Match", "Token F1", "BLEU", "chrF++", "TER",
            "ROUGE-1", "ROUGE-2", "ROUGE-L",
            "Score Expert (0–5)", "Commentaire Expert"]
    for ci, h in enumerate(hdrs, 1):
        cell = ws_exp.cell(1, ci, h)
        bg = "1F4E79" if ci<=4 else ("2E75B6" if ci<=12 else "C55A11")
        _header_style(cell, bg)

    guide = ("Guide notation expert :\n"
             "5 — Parfaite (natif MSA)\n4 — Très bonne\n"
             "3 — Acceptable\n2 — Partielle\n1 — Mauvaise\n0 — Hors sujet")
    ws_exp["M1"].comment = Comment(guide, "Zero-Shot Evaluation")

    for ri, row in enumerate(results, 2):
        mv = "✓" if row["exact_match"] else "✗"
        fc = "E2EFDA" if row["exact_match"] else "FCE4D6"
        dr = [row["id"], row["input_tun"], row["reference"], row["predicted"],
              mv, row["token_f1"], row["bleu"], row["chrf_pp"], row["ter"],
              row["rouge1"], row["rouge2"], row["rougeL"],
              row["expert_score"], row["expert_comment"]]
        for ci, val in enumerate(dr, 1):
            cell = ws_exp.cell(ri, ci, val)
            _data_style(cell, wrap=(ci in (2,3,4,14)))
            if ci in (2,3,4,5):
                cell.fill = PatternFill("solid", start_color=fc)

    nr = len(results) + 1
    for cl in ("F","J","K","L"):
        _pct_bar(ws_exp, cl, 2, nr)

    dv = DataValidation(type="whole", operator="between", formula1=0, formula2=5,
                        showErrorMessage=True, errorTitle="Invalide",
                        error="Entrez 0–5.")
    dv.sqref = f"M2:M{nr}"
    ws_exp.add_data_validation(dv)

    for col, w in {"A":6,"B":40,"C":40,"D":40,"E":12,"F":11,"G":9,
                   "H":10,"I":9,"J":10,"K":10,"L":10,"M":18,"N":40}.items():
        ws_exp.column_dimensions[col].width = w
    for r in range(2, nr+1):
        ws_exp.row_dimensions[r].height = 55

    # ── Onglet 3 : Échantillon 50 ─────────────────────────────
    ws_s = wb.create_sheet("🎲 Échantillon 50")
    ws_s.freeze_panes = "A2"
    idx50 = random.sample(range(len(results)), min(50, len(results)))
    rows50 = [results[i] for i in sorted(idx50)]
    for ci, h in enumerate(["ID","Texte Tunisien","Référence MSA",
                             f"Zéro-shot {model_display_name}",
                             "Token F1","BLEU","Score Expert (0–5)","Commentaire"],1):
        _header_style(ws_s.cell(1,ci,h))
    for ri, row in enumerate(rows50, 2):
        fc = "E2EFDA" if row["exact_match"] else "FFFFFF"
        for ci, val in enumerate([row["id"],row["input_tun"],row["reference"],
                                   row["predicted"],row["token_f1"],row["bleu"],
                                   row["expert_score"],row["expert_comment"]],1):
            cell = ws_s.cell(ri, ci, val)
            _data_style(cell, wrap=(ci in (2,3,4,8)))
            if ci in (2,3,4):
                cell.fill = PatternFill("solid", start_color=fc)
    for col, w in {"A":6,"B":42,"C":42,"D":42,"E":11,"F":9,"G":18,"H":40}.items():
        ws_s.column_dimensions[col].width = w
    for r in range(2, len(rows50)+2):
        ws_s.row_dimensions[r].height = 55

    dv2 = DataValidation(type="whole", operator="between", formula1=0, formula2=5,
                         showErrorMessage=True, errorTitle="Invalide",
                         error="Entrez 0–5.")
    dv2.sqref = f"G2:G{len(rows50)+1}"
    ws_s.add_data_validation(dv2)

    # ── Onglet 4 : Distribution ───────────────────────────────
    ws_d = wb.create_sheet("📈 Distribution")

    def _bucket(vals, metric_name, buckets, row_start):
        ws_d.cell(row_start, 1, metric_name).font = Font(bold=True, name="Arial", size=11)
        for c, h in enumerate(["Intervalle","Nombre","% du total"],1):
            _header_style(ws_d.cell(row_start+1, c))
            ws_d.cell(row_start+1, c, h)
        counts = [sum(1 for v in vals if lo <= v < hi) for lo, hi in buckets]
        for r, ((lo,hi),cnt) in enumerate(zip(buckets,counts), row_start+2):
            ws_d.cell(r,1,f"[{lo:.2f},{hi:.2f})")
            ws_d.cell(r,2,cnt)
            ws_d.cell(r,3, (f"={get_column_letter(2)}{r}/"
                             f"SUM({get_column_letter(2)}{row_start+2}:"
                             f"{get_column_letter(2)}{r+len(buckets)-1})"))
            ws_d.cell(r,3).number_format = "0.0%"

    bk = [(i/10,(i+1)/10) for i in range(10)]
    _bucket([r["token_f1"]      for r in results], "Token F1",      bk,  1)
    _bucket([r["bleu"]/100      for r in results], "BLEU (norm.)",  bk, 15)
    _bucket([r["chrf_pp"]/100   for r in results], "chrF++",        bk, 29)
    _bucket([r["rouge1"]        for r in results], "ROUGE-1",       bk, 43)
    for col, w in {"A":18,"B":12,"C":12}.items():
        ws_d.column_dimensions[col].width = w

    wb.save(out_path)
    print(f"  📊 Excel → {out_path}")

# ══════════════════════════════════════════════════════════════
# 5. BOUCLE PRINCIPALE — ÉVALUATION SÉQUENTIELLE DES 3 MODÈLES
# ══════════════════════════════════════════════════════════════
all_summaries = []

for model_cfg in MODELS:

    MODEL_ID     = model_cfg["model_id"]
    MODEL_NAME   = model_cfg["name"]
    SHORT_NAME   = model_cfg["short_name"]
    TRUST_REMOTE = model_cfg["trust_remote"]
    PROMPT_FN    = model_cfg["prompt_fn"]

    print("\n" + SEP)
    print(f"  🤖 MODÈLE : {MODEL_NAME}  ({MODEL_ID})")
    print(f"  Mode      : ZERO-SHOT (aucun adapteur LoRA)")
    print(SEP)

    # ── 5a. Quantification 4-bit ──────────────────────────────
    bnb_config = BitsAndBytesConfig(
        load_in_4bit=True,
        bnb_4bit_quant_type="nf4",
        bnb_4bit_compute_dtype=torch.bfloat16,
        bnb_4bit_use_double_quant=True,
    )

    # ── 5b. Chargement modèle de BASE uniquement (pas de LoRA) ─
    print(f"  📥 Chargement {MODEL_ID} ...")
    model = AutoModelForCausalLM.from_pretrained(
        MODEL_ID,
        quantization_config=bnb_config,
        device_map="auto",
        torch_dtype=torch.bfloat16,
        trust_remote_code=TRUST_REMOTE,
    )
    model.eval()
    model.config.use_cache = True

    # Fix dtype
    for _nm, _mod in model.named_modules():
        if hasattr(_mod, "weight") and _mod.weight is not None:
            if _mod.weight.dtype == torch.float32:
                _mod.weight.data = _mod.weight.data.to(torch.bfloat16)
        if hasattr(_mod, "bias") and _mod.bias is not None:
            if _mod.bias.dtype == torch.float32:
                _mod.bias.data = _mod.bias.data.to(torch.bfloat16)

    tokenizer = AutoTokenizer.from_pretrained(
        MODEL_ID,
        trust_remote_code=TRUST_REMOTE,
        padding_side="right",
    )
    if tokenizer.pad_token is None:
        tokenizer.pad_token     = tokenizer.eos_token
        tokenizer.pad_token_id  = tokenizer.eos_token_id

    print(f"  ✓ Modèle chargé — {sum(p.numel() for p in model.parameters())/1e9:.2f}B params\n")

    # ── 5c. Boucle d'inférence ────────────────────────────────
    print(f"  📊 Évaluation sur {len(test_data)} exemples ...")
    results_eval = []
    all_preds, all_refs = [], []
    N = len(test_data)

    for i, item in enumerate(test_data, 1):
        prompt = PROMPT_FN(item)
        inputs = tokenizer(prompt, return_tensors="pt")
        inputs = {k: v.to(model.device) for k, v in inputs.items()}

        with torch.no_grad():
            out = model.generate(
                input_ids=inputs["input_ids"].long(),
                attention_mask=inputs["attention_mask"].long(),
                max_new_tokens=MAX_NEW_TOKENS,
                do_sample=False,
                repetition_penalty=1.1,
                pad_token_id=tokenizer.eos_token_id,
                eos_token_id=tokenizer.eos_token_id,
            )
        new_tok = out[0][inputs["input_ids"].shape[1]:]
        pred = tokenizer.decode(new_tok, skip_special_tokens=True).strip()
        ref  = item["output"].strip()

        match   = is_exact_match(pred, ref)
        bleu_s  = sentence_bleu(pred, ref)
        ter_s   = sentence_ter(pred, ref)
        chrf_s  = sentence_chrf(pred, ref)
        rouge_s = sentence_rouge(pred, ref)
        tf1_s   = token_f1(pred, ref)

        all_preds.append(pred)
        all_refs.append(ref)

        results_eval.append({
            "id"             : i,
            "instruction"    : item["instruction"],
            "input_tun"      : item["input"],
            "reference"      : ref,
            "predicted"      : pred,
            "exact_match"    : match,
            "token_f1"       : round(tf1_s, 4),
            "bleu"           : round(bleu_s, 2),
            "chrf_pp"        : round(chrf_s, 2),
            "ter"            : round(ter_s, 2),
            "rouge1"         : round(rouge_s["rouge1"], 4),
            "rouge2"         : round(rouge_s["rouge2"], 4),
            "rougeL"         : round(rouge_s["rougeL"], 4),
            "expert_score"   : "",
            "expert_comment" : "",
        })

        if i <= 5 or i % 100 == 0:
            sym = "✓" if match else "✗"
            print(f"    Q{i:4d} {sym}  BLEU={bleu_s:.1f}  chrF++={chrf_s:.1f}"
                  f"  F1={tf1_s:.3f}")
            print(f"    Tunisien  : {item['input']}")
            print(f"    Référence : {ref}")
            print(f"    Prédit    : {pred[:120]}")

    # ── 5d. Métriques corpus ──────────────────────────────────
    c_bleu = _bleu_corp.corpus_score(all_preds, [all_refs])
    c_chrf = CHRF(word_order=2).corpus_score(all_preds, [all_refs])
    c_ter  = TER().corpus_score(all_preds, [all_refs])

    exact_matches = sum(r["exact_match"] for r in results_eval)
    exact_acc     = exact_matches / N
    avg_tf1       = float(np.mean([r["token_f1"] for r in results_eval]))
    avg_r1        = float(np.mean([r["rouge1"]   for r in results_eval]))
    avg_r2        = float(np.mean([r["rouge2"]   for r in results_eval]))
    avg_rL        = float(np.mean([r["rougeL"]   for r in results_eval]))

    summary = {
        "model"           : MODEL_ID,
        "mode"            : "zero-shot",
        "task"            : "Traduction dialecte tunisien → MSA",
        "n_test"          : N,
        "exact_match_acc" : round(exact_acc, 4),
        "avg_token_f1"    : round(avg_tf1, 4),
        "corpus_bleu"     : round(c_bleu.score, 2),
        "corpus_chrf_pp"  : round(c_chrf.score, 2),
        "corpus_ter"      : round(c_ter.score, 2),
        "avg_rouge1"      : round(avg_r1, 4),
        "avg_rouge2"      : round(avg_r2, 4),
        "avg_rougeL"      : round(avg_rL, 4),
    }
    all_summaries.append({**summary, "short_name": SHORT_NAME,
                          "display_name": MODEL_NAME})

    print(f"\n  {'─'*56}")
    print(f"  🎯 RÉSULTATS ZERO-SHOT — {MODEL_NAME}")
    print(f"  {'─'*56}")
    print(f"  Exact Match   : {exact_acc*100:.2f}%  ({exact_matches}/{N})")
    print(f"  Token F1      : {avg_tf1:.4f}")
    print(f"  BLEU (corpus) : {c_bleu.score:.2f}")
    print(f"  chrF++        : {c_chrf.score:.2f}")
    print(f"  TER           : {c_ter.score:.2f}  (↓ meilleur)")
    print(f"  ROUGE-1       : {avg_r1:.4f}")
    print(f"  ROUGE-2       : {avg_r2:.4f}")
    print(f"  ROUGE-L       : {avg_rL:.4f}")

    # ── 5e. Sauvegarde JSON ───────────────────────────────────
    eval_summary = {**summary, "predictions": results_eval}
    json_path = os.path.join(OUTPUT_DIR, f"baseline_zeroshot_{SHORT_NAME}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(eval_summary, f, ensure_ascii=False, indent=2)
    print(f"  💾 JSON  → {json_path}")

    # ── 5f. Export Excel ──────────────────────────────────────
    xlsx_path = os.path.join(OUTPUT_DIR, f"baseline_zeroshot_{SHORT_NAME}.xlsx")
    export_excel(results_eval, summary, MODEL_NAME, xlsx_path)

    # ── 5g. Libération mémoire GPU avant le prochain modèle ───
    del model, tokenizer
    gc.collect()
    torch.cuda.empty_cache()
    print(f"  ✓ GPU libéré pour le prochain modèle\n")

# ══════════════════════════════════════════════════════════════
# 6. TABLEAU COMPARATIF GLOBAL (JSON)
# ══════════════════════════════════════════════════════════════
comp_path = os.path.join(OUTPUT_DIR, "baseline_zeroshot_comparaison.json")
with open(comp_path, "w", encoding="utf-8") as f:
    json.dump(all_summaries, f, ensure_ascii=False, indent=2)

print("\n" + SEP)
print("  📊 COMPARAISON ZERO-SHOT — 3 MODÈLES")
print(SEP)
hdr = f"  {'Modèle':<22} {'BLEU':>8} {'chrF++':>8} {'TER':>8} {'F1':>8} {'Exact%':>8}"
print(hdr)
print("  " + "─"*62)
for s in all_summaries:
    print(f"  {s['display_name']:<22} "
          f"{s['corpus_bleu']:>8.2f} "
          f"{s['corpus_chrf_pp']:>8.2f} "
          f"{s['corpus_ter']:>8.2f} "
          f"{s['avg_token_f1']:>8.4f} "
          f"{s['exact_match_acc']*100:>7.2f}%")
print(SEP)
print(f"  💾 Comparaison JSON → {comp_path}")
print(f"  📁 Dossier sortie   → {OUTPUT_DIR}")
print(SEP)
print("  ✅ ÉVALUATION ZERO-SHOT BASELINE TERMINÉE!")
print(SEP)
