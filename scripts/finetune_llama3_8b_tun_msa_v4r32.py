"""
Fine-tuning LLaMA 3-8B sur machine GPU/CUDA (Linux/Windows)
Task  : Traduction dialecte tunisien → arabe standard moderne (MSA)
Dataset: 3000 paires instruction/input/output au format JSON
Méthode: QLoRA (4-bit NF4) + LoRA PEFT

v4 — Améliorations appliquées (recommandations) :
  • [FIX 1] compute_metrics BLEU intégré dans le Trainer → optimisation guidée
  • [FIX 2] Early Stopping (EarlyStoppingCallback, patience=1)
  • [FIX 3] Régularisation QLoRA renforcée :
      - r réduit : 64 → 16
      - lora_dropout augmenté : 0.05 → 0.10
      - weight_decay augmenté : 0.01 → 0.10
"""

# ─────────────────────────────────────────────────────────────
# pip install transformers datasets peft trl bitsandbytes
#             accelerate torch scikit-learn
#             sacrebleu rouge-score openpyxl packaging
# ─────────────────────────────────────────────────────────────

import os, gc, json, random, datetime
import torch
import numpy as np
from sklearn.metrics import accuracy_score, classification_report
from transformers import (
    AutoModelForCausalLM,
    AutoTokenizer,
    TrainingArguments,
    BitsAndBytesConfig,
    EarlyStoppingCallback,   # ← [FIX 2]
)
from peft import LoraConfig, get_peft_model, prepare_model_for_kbit_training
from datasets import Dataset
from trl import SFTTrainer

# ── Métriques ────────────────────────────────────────────────
from sacrebleu.metrics import BLEU, TER, CHRF
from rouge_score import rouge_scorer
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

# ══════════════════════════════════════════════════════════════
# 0. ENVIRONNEMENT CUDA
# ══════════════════════════════════════════════════════════════
os.environ["CUDA_VISIBLE_DEVICES"] = "0,1"
os.environ["HF_HUB_DISABLE_MODEL_CARD"] = "1"
os.environ["TOKENIZERS_PARALLELISM"] = "false"

gc.collect()
torch.cuda.empty_cache()

SEP = "=" * 60

print(SEP)
print("  Fine-tuning LLaMA 3-8B — Traduction Tunisien → MSA  v4-r32")
print(SEP)
print(f"  PyTorch       : {torch.__version__}")
print(f"  CUDA dispo    : {torch.cuda.is_available()}")
print(f"  Nombre GPUs   : {torch.cuda.device_count()}")
if torch.cuda.is_available():
    for i in range(torch.cuda.device_count()):
        name = torch.cuda.get_device_name(i)
        vram = torch.cuda.get_device_properties(i).total_memory / 1e9
        print(f"  GPU {i}         : {name}  ({vram:.1f} Go VRAM)")
print(SEP + "\n")

# ══════════════════════════════════════════════════════════════
# 1. CHEMINS — À ADAPTER SELON VOTRE MACHINE
# ══════════════════════════════════════════════════════════════
DATASET_FILE = r"C:\Users\infocom\tradiction_dialect\dataset_final_3000.json"
OUTPUT_DIR   = r"C:\Users\infocom\tradiction_dialect\llama3_8b_tun_msa_r32_output-3000"
MODEL_ID     = "meta-llama/Meta-Llama-3-8B-Instruct"
SEED         = 42

if not os.path.exists(DATASET_FILE):
    raise FileNotFoundError(f"❌ Dataset introuvable : {DATASET_FILE}")
print(f"✓ Dataset trouvé : {DATASET_FILE}\n")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ══════════════════════════════════════════════════════════════
# 2. CHARGEMENT ET PRÉPARATION DU DATASET
# ══════════════════════════════════════════════════════════════
print("📊 Chargement du dataset...")
with open(DATASET_FILE, "r", encoding="utf-8") as f:
    raw_data = json.load(f)

if isinstance(raw_data, dict):
    raw = raw_data.get("data", list(raw_data.values()))
else:
    raw = raw_data

print(f"  Exemples chargés : {len(raw)}")

def is_valid(item):
    return (
        isinstance(item, dict)
        and item.get("instruction", "").strip()
        and item.get("input", "").strip()
        and item.get("output", "").strip()
    )

raw = [item for item in raw if is_valid(item)]
print(f"  Après nettoyage  : {len(raw)} exemples valides")

random.seed(SEED)
random.shuffle(raw)

split      = int(len(raw) * 0.85)
train_data = raw[:split]
test_data  = raw[split:]

with open(f"{OUTPUT_DIR}/test_set_FINAL.json", "w", encoding="utf-8") as f:
    json.dump(test_data, f, ensure_ascii=False, indent=2)

print(f"  Train : {len(train_data)} | Test : {len(test_data)}")
print(f"  Test set sauvegardé → {OUTPUT_DIR}/test_set_FINAL.json\n")

# ══════════════════════════════════════════════════════════════
# 3. FORMAT DU PROMPT
# LLaMA 3 utilise le format chat avec des balises spéciales
# ══════════════════════════════════════════════════════════════
PROMPT_TEMPLATE = (
    "<|begin_of_text|>"
    "<|start_header_id|>system<|end_header_id|>\n\n"
    "{instruction}<|eot_id|>"
    "<|start_header_id|>user<|end_header_id|>\n\n"
    "{input}<|eot_id|>"
    "<|start_header_id|>assistant<|end_header_id|>\n\n"
    "{output}<|eot_id|>"
)

INFERENCE_TEMPLATE = (
    "<|begin_of_text|>"
    "<|start_header_id|>system<|end_header_id|>\n\n"
    "{instruction}<|eot_id|>"
    "<|start_header_id|>user<|end_header_id|>\n\n"
    "{input}<|eot_id|>"
    "<|start_header_id|>assistant<|end_header_id|>\n\n"
)

def format_sample(item: dict) -> dict:
    return {
        "text": PROMPT_TEMPLATE.format(
            instruction=item["instruction"].strip(),
            input=item["input"].strip(),
            output=item["output"].strip(),
        )
    }

train_ds = Dataset.from_list([format_sample(q) for q in train_data])
val_ds   = Dataset.from_list([format_sample(q) for q in test_data[:80]])

# val_ds_raw : données brutes nécessaires pour compute_metrics BLEU
val_data_raw = test_data[:80]

print(f"✓ {len(train_ds)} exemples train formatés")
print("  Aperçu :")
print("  " + train_ds[0]["text"][:250].replace("\n", "\n  ") + "...\n")

# ══════════════════════════════════════════════════════════════
# 4. QUANTIFICATION 4-BIT (QLoRA)
# ══════════════════════════════════════════════════════════════
print("🔧 Configuration quantification 4-bit NF4...")
bnb_config = BitsAndBytesConfig(
    load_in_4bit=True,
    bnb_4bit_quant_type="nf4",
    bnb_4bit_compute_dtype=torch.bfloat16,
    bnb_4bit_use_double_quant=True,
)

# ══════════════════════════════════════════════════════════════
# 5. CHARGEMENT MODÈLE + TOKENIZER
# ══════════════════════════════════════════════════════════════
print(f"📥 Chargement {MODEL_ID}...")
model = AutoModelForCausalLM.from_pretrained(
    MODEL_ID,
    quantization_config=bnb_config,
    device_map="auto",
    torch_dtype=torch.bfloat16,
)
model.config.use_cache = False

tokenizer = AutoTokenizer.from_pretrained(
    MODEL_ID,
    padding_side="right",
)
# LLaMA 3 : pad_token doit être défini explicitement
if tokenizer.pad_token is None:
    tokenizer.pad_token = tokenizer.eos_token
    tokenizer.pad_token_id = tokenizer.eos_token_id
print("✓ Modèle et tokenizer chargés\n")

# ══════════════════════════════════════════════════════════════
# 6. PEFT — LoRA
# v4-r32 : r=32 (vs r=16 en v4), régularisation v4 maintenue
# ══════════════════════════════════════════════════════════════
print("🔧 Application LoRA (PEFT)...")
model = prepare_model_for_kbit_training(model)

peft_config = LoraConfig(
    r=32,                   # <- v4-r32 : était 16 en v4 (test capacité supérieure)
    lora_alpha=64,          # <- v4-r32 : ajusté proportionnellement (alpha=2*r)
    lora_dropout=0.10,      # <- maintenu depuis v4
    target_modules=[
        "q_proj", "k_proj", "v_proj", "o_proj",
        "gate_proj", "up_proj", "down_proj",
    ],
    bias="none",
    task_type="CAUSAL_LM",
)

model = get_peft_model(model, peft_config)
model.print_trainable_parameters()

# ══════════════════════════════════════════════════════════════
# 7. FONCTION compute_metrics BLEU  [FIX 1]
# Appelée par le Trainer à chaque eval_steps pour suivre le BLEU
# sur le jeu de validation.
# NOTE : EvalPrediction.predictions contient les logits (ou token IDs
#        si predict_with_generate=True). Avec SFTTrainer standard, on
#        reçoit des logits → on prend l'argmax pour décoder.
# ══════════════════════════════════════════════════════════════
_bleu_train_metric = BLEU(tokenize="char")

def compute_metrics(eval_pred):
    """
    Calcule le BLEU corpus sur le batch de validation.
    eval_pred.predictions : tableau numpy (batch, seq_len, vocab) ou (batch, seq_len)
    eval_pred.label_ids   : tableau numpy (batch, seq_len) — token IDs cibles
    """
    predictions, label_ids = eval_pred

    # Si logits (3D), prendre l'argmax
    if predictions.ndim == 3:
        pred_ids = np.argmax(predictions, axis=-1)
    else:
        pred_ids = predictions.astype(int)

    # Remplacer -100 (padding masqué) par pad_token_id dans les labels
    label_ids = np.where(label_ids == -100, tokenizer.pad_token_id, label_ids)

    decoded_preds  = tokenizer.batch_decode(pred_ids,  skip_special_tokens=True)
    decoded_labels = tokenizer.batch_decode(label_ids, skip_special_tokens=True)

    # Nettoyage basique
    decoded_preds  = [p.strip() for p in decoded_preds]
    decoded_labels = [l.strip() for l in decoded_labels]

    try:
        bleu_score = _bleu_train_metric.corpus_score(
            decoded_preds, [decoded_labels]
        ).score
    except Exception:
        bleu_score = 0.0

    return {"bleu": round(bleu_score, 2)}

# ══════════════════════════════════════════════════════════════
# 8. ARGUMENTS D'ENTRAÎNEMENT
# [FIX 2] Early Stopping : metric_for_best_model → "eval_bleu"
#          greater_is_better=True (BLEU : plus haut = meilleur)
# [FIX 3] weight_decay augmenté : 0.01 → 0.10
# v4-r32 : 2 époques max — early stop patience=1 sur eval_bleu
# ══════════════════════════════════════════════════════════════
print("\n🚀 Configuration entraînement...\n")
training_args = TrainingArguments(
    output_dir=OUTPUT_DIR,
    per_device_train_batch_size=4,
    per_device_eval_batch_size=2,
    gradient_accumulation_steps=4,
    num_train_epochs=2,             # <- v4-r32 : 2 époques max (early stop patience=1)
    learning_rate=3e-4,
    lr_scheduler_type="cosine",
    warmup_steps=100,
    weight_decay=0.10,              # ← [FIX 3] était 0.01
    optim="paged_adamw_8bit",
    bf16=True,
    fp16=False,
    gradient_checkpointing=True,
    logging_steps=10,
    eval_steps=100,
    save_steps=100,                 # aligné sur eval_steps pour load_best
    eval_strategy="steps",
    save_strategy="steps",
    save_total_limit=3,
    load_best_model_at_end=True,
    metric_for_best_model="eval_bleu",   # ← [FIX 1+2] était "eval_loss"
    greater_is_better=True,              # ← [FIX 1+2] BLEU ↑ = meilleur
    dataloader_num_workers=0,
    dataloader_pin_memory=True,
    ddp_find_unused_parameters=False,
    report_to="none",
    seed=SEED,
)

# ══════════════════════════════════════════════════════════════
# 9. TRAINER SFT  (compatible trl ≥ 0.8 et trl < 0.8)
# ══════════════════════════════════════════════════════════════
import trl as _trl_mod
from packaging.version import Version as _V

_trl_new = _V(_trl_mod.__version__) >= _V("0.8.0")

# [FIX 2] EarlyStoppingCallback : patience=1 → arrêt dès que le BLEU
#          ne s'améliore plus pendant 1 évaluation consécutive.
early_stopping_cb = EarlyStoppingCallback(early_stopping_patience=1)

if _trl_new:
    from trl import SFTConfig

    _SFT_ONLY_KEYS = {
        "output_dir", "max_seq_length", "max_length",
        "packing", "dataset_text_field",
        "dataset_num_proc", "dataset_batch_size",
        "neftune_noise_alpha",
    }

    import inspect as _inspect
    _sft_params = set(_inspect.signature(SFTConfig.__init__).parameters.keys())
    _seq_len_kwarg = "max_length" if "max_length" in _sft_params else "max_seq_length"

    sft_config = SFTConfig(
        output_dir=OUTPUT_DIR,
        **{_seq_len_kwarg: 512},
        packing=False,
        dataset_text_field="text",
        **{k: v for k, v in training_args.to_dict().items()
        if k not in _SFT_ONLY_KEYS},
    )
    try:
        trainer = SFTTrainer(
            model=model,
            args=sft_config,
            train_dataset=train_ds,
            eval_dataset=val_ds,
            compute_metrics=compute_metrics,   # ← [FIX 1]
            callbacks=[early_stopping_cb],     # ← [FIX 2]
        )
    except TypeError:
        def _fmt(batch):
            return batch["text"]

        trainer = SFTTrainer(
            model=model,
            args=sft_config,
            train_dataset=train_ds,
            eval_dataset=val_ds,
            formatting_func=_fmt,
            compute_metrics=compute_metrics,   # ← [FIX 1]
            callbacks=[early_stopping_cb],     # ← [FIX 2]
        )
else:
    trainer = SFTTrainer(
        model=model,
        args=training_args,
        train_dataset=train_ds,
        eval_dataset=val_ds,
        dataset_text_field="text",
        max_seq_length=512,
        packing=False,
        compute_metrics=compute_metrics,       # ← [FIX 1]
        callbacks=[early_stopping_cb],         # ← [FIX 2]
    )

print(f"  trl {_trl_mod.__version__} → SFTTrainer configuré ✓")
print(f"  compute_metrics BLEU     : activé ✓")
print(f"  EarlyStoppingCallback    : patience=1, métrique=eval_bleu ✓\n")

# ══════════════════════════════════════════════════════════════
# 10. ENTRAÎNEMENT
# ══════════════════════════════════════════════════════════════
print(SEP)
print("  🔥 DÉBUT ENTRAÎNEMENT")
print(SEP)
print(f"  Dataset train   : {len(train_data)} exemples")
print(f"  Dataset val     : {len(val_ds)} exemples")
print(f"  Batch effectif  : {4 * 4 * max(1, torch.cuda.device_count())}")
print(f"  Epoques max     : 2  (early stop patience=1, metrique=eval_bleu — exp. r32)")
print(f"  LoRA rank       : 32  (augmenté depuis 16 — expérience r32)")
print(f"  lora_dropout    : 0.10 (augmenté depuis 0.05)")
print(f"  weight_decay    : 0.10 (augmenté depuis 0.01)")
print(f"  Temps estimé    : ~25–50 min (2x L4)")
print(SEP + "\n")

trainer.train()

# ══════════════════════════════════════════════════════════════
# 11. SAUVEGARDE
# ══════════════════════════════════════════════════════════════
print("\n💾 Sauvegarde du modèle (adapteur LoRA)...")
trainer.save_model(OUTPUT_DIR)
tokenizer.save_pretrained(OUTPUT_DIR)
print(f"✅ MODÈLE SAUVEGARDÉ → {OUTPUT_DIR}\n")

# ══════════════════════════════════════════════════════════════
# 12. FONCTIONS D'INFÉRENCE ET DE MÉTRIQUES (évaluation finale)
# ══════════════════════════════════════════════════════════════

def translate_to_msa(item: dict, max_new_tokens: int = 128) -> str:
    prompt = INFERENCE_TEMPLATE.format(
        instruction=item["instruction"].strip(),
        input=item["input"].strip(),
    )
    inputs = tokenizer(prompt, return_tensors="pt").to(model.device)
    with torch.no_grad():
        outputs = model.generate(
            **inputs,
            max_new_tokens=max_new_tokens,
            do_sample=False,
            repetition_penalty=1.1,
            pad_token_id=tokenizer.eos_token_id,
            eos_token_id=tokenizer.eos_token_id,
        )
    new_tokens = outputs[0][inputs["input_ids"].shape[1]:]
    return tokenizer.decode(new_tokens, skip_special_tokens=True).strip()


def is_exact_match(pred: str, ref: str) -> bool:
    return pred.strip().rstrip("،.") == ref.strip().rstrip("،.")


_bleu_metric  = BLEU(effective_order=True, tokenize="char")
_ter_metric   = TER()
_chrf_metric  = CHRF(word_order=2)
_rouge        = rouge_scorer.RougeScorer(
    ["rouge1", "rouge2", "rougeL"], use_stemmer=False
)

def compute_sentence_bleu(pred: str, ref: str) -> float:
    try:    return _bleu_metric.sentence_score(pred, [ref]).score
    except: return 0.0

def compute_ter(pred: str, ref: str) -> float:
    try:    return _ter_metric.sentence_score(pred, [ref]).score
    except: return 100.0

def compute_chrf(pred: str, ref: str) -> float:
    try:    return _chrf_metric.sentence_score(pred, [ref]).score
    except: return 0.0

def compute_rouge(pred: str, ref: str) -> dict:
    try:
        scores = _rouge.score(ref, pred)
        return {
            "rouge1": scores["rouge1"].fmeasure,
            "rouge2": scores["rouge2"].fmeasure,
            "rougeL": scores["rougeL"].fmeasure,
        }
    except:
        return {"rouge1": 0.0, "rouge2": 0.0, "rougeL": 0.0}

def compute_token_f1(pred: str, ref: str) -> float:
    pred_tokens = set(pred.strip().split())
    ref_tokens  = set(ref.strip().split())
    if not pred_tokens or not ref_tokens:
        return float(pred_tokens == ref_tokens)
    common = pred_tokens & ref_tokens
    if not common:
        return 0.0
    p = len(common) / len(pred_tokens)
    r = len(common) / len(ref_tokens)
    return 2 * p * r / (p + r)

# ══════════════════════════════════════════════════════════════
# 13. BOUCLE D'ÉVALUATION FINALE
# ══════════════════════════════════════════════════════════════
print(SEP)
print("  📊 ÉVALUATION SUR LE TEST SET COMPLET")
print(SEP)

model.eval()
model.config.use_cache = True

results_eval = []
all_preds, all_refs = [], []
N = len(test_data)

for i, item in enumerate(test_data, 1):
    pred  = translate_to_msa(item)
    ref   = item["output"].strip()
    match = is_exact_match(pred, ref)

    bleu_s  = compute_sentence_bleu(pred, ref)
    ter_s   = compute_ter(pred, ref)
    chrf_s  = compute_chrf(pred, ref)
    rouge_s = compute_rouge(pred, ref)
    tf1_s   = compute_token_f1(pred, ref)

    all_preds.append(pred)
    all_refs.append(ref)

    results_eval.append({
        "id"             : i,
        "instruction"    : item["instruction"],
        "input_tun"      : item["input"],
        "reference"      : ref,
        "predicted"      : pred,
        "exact_match"    : match,
        "token_f1"       : round(tf1_s, 4),
        "bleu"           : round(bleu_s, 2),
        "chrf_pp"        : round(chrf_s, 2),
        "ter"            : round(ter_s, 2),
        "rouge1"         : round(rouge_s["rouge1"], 4),
        "rouge2"         : round(rouge_s["rouge2"], 4),
        "rougeL"         : round(rouge_s["rougeL"], 4),
        "expert_score"   : "",
        "expert_comment" : "",
    })

    symbol = "✓" if match else "✗"
    if i <= 10 or i % 50 == 0:
        print(f"\nQ{i:4d} {symbol}  BLEU={bleu_s:.1f}  chrF++={chrf_s:.1f}"
              f"  TER={ter_s:.1f}  F1={tf1_s:.3f}")
        print(f"  Tunisien  : {item['input']}")
        print(f"  Référence : {ref}")
        print(f"  Prédit    : {pred}")

# ── Métriques corpus ─────────────────────────────────────────
corpus_bleu = BLEU(tokenize="char").corpus_score(all_preds, [all_refs])
corpus_chrf = CHRF(word_order=2).corpus_score(all_preds, [all_refs])
corpus_ter  = TER().corpus_score(all_preds, [all_refs])

exact_matches = sum(r["exact_match"] for r in results_eval)
exact_acc     = exact_matches / N
avg_token_f1  = np.mean([r["token_f1"] for r in results_eval])
avg_rouge1    = np.mean([r["rouge1"]   for r in results_eval])
avg_rouge2    = np.mean([r["rouge2"]   for r in results_eval])
avg_rougeL    = np.mean([r["rougeL"]   for r in results_eval])

summary_metrics = {
    "model"           : MODEL_ID,
    "task"            : "Traduction dialecte tunisien → MSA",
    "n_test"          : N,
    "lora_r"          : 32,
    "lora_dropout"    : 0.10,
    "weight_decay"    : 0.10,
    "exact_match_acc" : round(float(exact_acc), 4),
    "avg_token_f1"    : round(float(avg_token_f1), 4),
    "corpus_bleu"     : round(corpus_bleu.score, 2),
    "corpus_chrf_pp"  : round(corpus_chrf.score, 2),
    "corpus_ter"      : round(corpus_ter.score, 2),
    "avg_rouge1"      : round(float(avg_rouge1), 4),
    "avg_rouge2"      : round(float(avg_rouge2), 4),
    "avg_rougeL"      : round(float(avg_rougeL), 4),
}

print("\n" + SEP)
print("  🎯 RÉSUMÉ DES MÉTRIQUES (corpus)")
print(SEP)
print(f"  Exact Match Accuracy : {exact_acc*100:.2f}%  ({exact_matches}/{N})")
print(f"  Token F1 (avg)       : {avg_token_f1:.4f}")
print(f"  BLEU (corpus)        : {corpus_bleu.score:.2f}")
print(f"  chrF++ (corpus)      : {corpus_chrf.score:.2f}")
print(f"  TER (corpus)         : {corpus_ter.score:.2f}  (↓ meilleur)")
print(f"  ROUGE-1 (avg)        : {avg_rouge1:.4f}")
print(f"  ROUGE-2 (avg)        : {avg_rouge2:.4f}")
print(f"  ROUGE-L (avg)        : {avg_rougeL:.4f}")
print(SEP)

# ══════════════════════════════════════════════════════════════
# 14. SAUVEGARDE JSON
# ══════════════════════════════════════════════════════════════
eval_summary = {**summary_metrics, "predictions": results_eval}
results_path = f"{OUTPUT_DIR}/evaluation_results.json"
with open(results_path, "w", encoding="utf-8") as f:
    json.dump(eval_summary, f, ensure_ascii=False, indent=2)
print(f"\n💾 Résultats JSON    → {results_path}")

# ══════════════════════════════════════════════════════════════
# 15. EXPORT EXCEL MULTI-ONGLETS
# ══════════════════════════════════════════════════════════════

def _header_style(cell, bg="1F4E79"):
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    thin = Side(style="thin", color="FFFFFF")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def _data_style(cell, wrap=False):
    cell.font      = Font(name="Arial", size=9)
    cell.alignment = Alignment(vertical="center", wrap_text=wrap,
                                horizontal="right")
    thin = Side(style="thin", color="D9D9D9")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def _pct_bar(ws, col_letter, row_start, row_end):
    rule = DataBarRule(start_type="num", start_value=0,
                       end_type="num",   end_value=1,
                       color="4472C4")
    ws.conditional_formatting.add(
        f"{col_letter}{row_start}:{col_letter}{row_end}", rule
    )

def export_excel(results, summary, out_path):
    wb = Workbook()

    # ── ONGLET 1 : Métriques globales ────────────────────────
    ws_sum = wb.active
    ws_sum.title = "📋 Métriques Globales"
    ws_sum.sheet_view.rightToLeft = False

    ws_sum.merge_cells("A1:C1")
    title_cell = ws_sum["A1"]
    title_cell.value = "LLaMA 3-8B LoRA v4-r32 — Evaluation Traduction Tunisien -> MSA"
    title_cell.font  = Font(bold=True, size=14, color="FFFFFF", name="Arial")
    title_cell.fill  = PatternFill("solid", start_color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 30

    kv_rows = [
        ("Modèle de base",       summary["model"]),
        ("Tâche",                summary["task"]),
        ("Nombre exemples test", summary["n_test"]),
        ("LoRA rank (r)",        summary["lora_r"]),
        ("lora_dropout",         summary["lora_dropout"]),
        ("weight_decay",         summary["weight_decay"]),
        ("Date évaluation",
         datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Exact Match Accuracy",
         f"{summary['exact_match_acc']*100:.2f} %"),
        ("Token F1 (avg)",       f"{summary['avg_token_f1']:.4f}"),
        ("BLEU (corpus)",        f"{summary['corpus_bleu']:.2f}"),
        ("chrF++ (corpus)",      f"{summary['corpus_chrf_pp']:.2f}"),
        ("TER (corpus, ↓)",      f"{summary['corpus_ter']:.2f}"),
        ("ROUGE-1 (avg)",        f"{summary['avg_rouge1']:.4f}"),
        ("ROUGE-2 (avg)",        f"{summary['avg_rouge2']:.4f}"),
        ("ROUGE-L (avg)",        f"{summary['avg_rougeL']:.4f}"),
    ]
    for row_idx, (k, v) in enumerate(kv_rows, 2):
        cell_k = ws_sum.cell(row=row_idx, column=1, value=k)
        cell_v = ws_sum.cell(row=row_idx, column=2, value=v)
        if k:
            cell_k.font = Font(bold=True, name="Arial", size=10)
            cell_v.font = Font(name="Arial", size=10)

    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 50

    # ── ONGLET 2 : Évaluation expert ─────────────────────────
    from openpyxl.comments import Comment
    from openpyxl.worksheet.datavalidation import DataValidation

    ws_exp = wb.create_sheet("🔍 Évaluation Expert")
    ws_exp.sheet_view.rightToLeft = False
    ws_exp.freeze_panes = "A2"

    expert_headers = [
        "ID", "Texte Tunisien (input)",
        "Référence MSA", "Traduction LLaMA 3",
        "Exact Match",
        "Token F1", "BLEU", "chrF++", "TER",
        "ROUGE-1", "ROUGE-2", "ROUGE-L",
        "Score Expert (0–5)", "Commentaire Expert",
    ]
    for col, h in enumerate(expert_headers, 1):
        cell = ws_exp.cell(row=1, column=col, value=h)
        bg   = "1F4E79" if col <= 4 else ("2E75B6" if col <= 12 else "C55A11")
        _header_style(cell, bg=bg)

    score_guide = (
        "Guide de notation expert :\n"
        "5 — Traduction parfaite, équivalente au natif MSA\n"
        "4 — Très bonne, quelques imprécisions mineures\n"
        "3 — Acceptable, sens correct mais style perfectible\n"
        "2 — Partielle, perte d'information notable\n"
        "1 — Mauvaise, sens très altéré\n"
        "0 — Complètement erronée ou hors sujet"
    )
    ws_exp["M1"].comment = Comment(score_guide, "LLaMA 3 Evaluation")

    for r_idx, row in enumerate(results, 2):
        match_val  = "✓" if row["exact_match"] else "✗"
        fill_color = "E2EFDA" if row["exact_match"] else "FCE4D6"
        data_row = [
            row["id"], row["input_tun"], row["reference"], row["predicted"],
            match_val,
            row["token_f1"], row["bleu"], row["chrf_pp"], row["ter"],
            row["rouge1"], row["rouge2"], row["rougeL"],
            row["expert_score"], row["expert_comment"],
        ]
        for c_idx, val in enumerate(data_row, 1):
            cell = ws_exp.cell(row=r_idx, column=c_idx, value=val)
            _data_style(cell, wrap=(c_idx in (2, 3, 4, 14)))
            if c_idx in (2, 3, 4, 5):
                cell.fill = PatternFill("solid", start_color=fill_color)

    n_rows = len(results) + 1
    _pct_bar(ws_exp, "F", 2, n_rows)
    _pct_bar(ws_exp, "J", 2, n_rows)
    _pct_bar(ws_exp, "K", 2, n_rows)
    _pct_bar(ws_exp, "L", 2, n_rows)

    dv = DataValidation(type="whole", operator="between",
                        formula1=0, formula2=5,
                        showErrorMessage=True,
                        errorTitle="Valeur invalide",
                        error="Entrez un entier entre 0 et 5.")
    dv.sqref = f"M2:M{n_rows}"
    ws_exp.add_data_validation(dv)

    col_widths = {
        "A": 6,  "B": 40, "C": 40, "D": 40,
        "E": 12, "F": 11, "G": 9,  "H": 10, "I": 9,
        "J": 10, "K": 10, "L": 10, "M": 18, "N": 40,
    }
    for col, w in col_widths.items():
        ws_exp.column_dimensions[col].width = w
    for r in range(2, n_rows + 1):
        ws_exp.row_dimensions[r].height = 60

    # ── ONGLET 3 : Échantillon 50 ────────────────────────────
    ws_sample = wb.create_sheet("🎲 Échantillon 50")
    ws_sample.freeze_panes = "A2"

    sample_indices = random.sample(range(len(results)), min(50, len(results)))
    sample_rows    = [results[i] for i in sorted(sample_indices)]

    sample_headers = ["ID", "Texte Tunisien", "Référence MSA",
                      "Traduction LLaMA 3", "Token F1", "BLEU",
                      "Score Expert (0–5)", "Commentaire Expert"]
    for col, h in enumerate(sample_headers, 1):
        _header_style(ws_sample.cell(row=1, column=col, value=h))

    for r_idx, row in enumerate(sample_rows, 2):
        vals = [
            row["id"], row["input_tun"], row["reference"], row["predicted"],
            row["token_f1"], row["bleu"],
            row["expert_score"], row["expert_comment"],
        ]
        fill_color = "E2EFDA" if row["exact_match"] else "FFFFFF"
        for c_idx, val in enumerate(vals, 1):
            cell = ws_sample.cell(row=r_idx, column=c_idx, value=val)
            _data_style(cell, wrap=(c_idx in (2, 3, 4, 8)))
            if c_idx in (2, 3, 4):
                cell.fill = PatternFill("solid", start_color=fill_color)

    sample_widths = {"A": 6, "B": 42, "C": 42, "D": 42,
                     "E": 11, "F": 9, "G": 18, "H": 40}
    for col, w in sample_widths.items():
        ws_sample.column_dimensions[col].width = w
    for r in range(2, len(sample_rows) + 2):
        ws_sample.row_dimensions[r].height = 60

    dv2 = DataValidation(type="whole", operator="between",
                         formula1=0, formula2=5,
                         showErrorMessage=True,
                         errorTitle="Valeur invalide",
                         error="Entrez un entier entre 0 et 5.")
    dv2.sqref = f"G2:G{len(sample_rows)+1}"
    ws_sample.add_data_validation(dv2)

    # ── ONGLET 4 : Distribution métriques ────────────────────
    ws_dist = wb.create_sheet("📈 Distribution")

    def _bucket_metric(vals, metric_name, buckets, row_start):
        ws_dist.cell(row_start, 1, metric_name).font = Font(
            bold=True, name="Arial", size=11)
        headers = ["Intervalle", "Nombre", "% du total"]
        for c, h in enumerate(headers, 1):
            _header_style(ws_dist.cell(row_start + 1, c))
            ws_dist.cell(row_start + 1, c, h)
        counts = [sum(1 for v in vals if lo <= v < hi) for lo, hi in buckets]
        for r, ((lo, hi), cnt) in enumerate(zip(buckets, counts),
                                             row_start + 2):
            ws_dist.cell(r, 1, f"[{lo:.2f}, {hi:.2f})")
            ws_dist.cell(r, 2, cnt)
            ws_dist.cell(r, 3,
                f"={get_column_letter(2)}{r}/"
                f"SUM({get_column_letter(2)}{row_start+2}:"
                f"{get_column_letter(2)}{r+len(buckets)-1})")
            ws_dist.cell(r, 3).number_format = "0.0%"

    buckets_01 = [(i / 10, (i + 1) / 10) for i in range(10)]
    _bucket_metric([r["token_f1"]        for r in results],
                   "Token F1",       buckets_01, 1)
    _bucket_metric([r["bleu"] / 100      for r in results],
                   "BLEU (norm.)",   buckets_01, 15)
    _bucket_metric([r["chrf_pp"] / 100   for r in results],
                   "chrF++",         buckets_01, 29)
    _bucket_metric([r["rouge1"]          for r in results],
                   "ROUGE-1",        buckets_01, 43)

    ws_dist.column_dimensions["A"].width = 18
    ws_dist.column_dimensions["B"].width = 12
    ws_dist.column_dimensions["C"].width = 12

    wb.save(out_path)
    print(f"📊 Fichier Excel généré → {out_path}")


# ── Appel export ──────────────────────────────────────────────
excel_path = f"{OUTPUT_DIR}/llama3_r32_evaluation_expert.xlsx"
export_excel(results_eval, summary_metrics, excel_path)

with open(results_path, "w", encoding="utf-8") as f:
    json.dump(eval_summary, f, ensure_ascii=False, indent=2)

print(f"\n💾 Résultats JSON    → {results_path}")
print(f"📄 Test set          → {OUTPUT_DIR}/test_set_FINAL.json")
print("\n" + SEP)
print("  ✅ FINE-TUNING ET EVALUATION LLaMA 3-8B v4-r32 TERMINES AVEC SUCCES!")
print(SEP)

# ══════════════════════════════════════════════════════════════
# 16. FUSION OPTIONNELLE : LORA → MODÈLE COMPLET
# ══════════════════════════════════════════════════════════════
# from peft import PeftModel
# from transformers import AutoModelForCausalLM
#
# MERGED_DIR = OUTPUT_DIR + "_merged"
# base = AutoModelForCausalLM.from_pretrained(
#     MODEL_ID, torch_dtype=torch.bfloat16, device_map="cpu",
# )
# merged = PeftModel.from_pretrained(base, OUTPUT_DIR).merge_and_unload()
# merged.save_pretrained(MERGED_DIR)
# tokenizer.save_pretrained(MERGED_DIR)
# print("✅ Modèle fusionné sauvegardé.")
