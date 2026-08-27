"""
Évaluation autonome — ALLaM-7B fine-tuné (adapteur LoRA)
Task  : Traduction dialecte tunisien → arabe standard moderne (MSA)
Charge le modèle de base + l'adapteur LoRA sauvegardé, puis évalue
sur le test set complet (test_set_FINAL.json).

Fix appliqué : cast explicite en bfloat16 avant generate() pour éviter
  RuntimeError: expected scalar type BFloat16 but found Float

pip install transformers datasets peft bitsandbytes accelerate torch
            sacrebleu rouge-score openpyxl packaging
"""

import os, gc, json, random, datetime
import torch
import numpy as np

from transformers import AutoModelForCausalLM, AutoTokenizer, BitsAndBytesConfig
from peft import PeftModel

from sacrebleu.metrics import BLEU, TER, CHRF
from rouge_score import rouge_scorer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation

# ══════════════════════════════════════════════════════════════
# 0. CHEMINS — À ADAPTER SELON VOTRE MACHINE
# ══════════════════════════════════════════════════════════════
MODEL_ID     = "humain-ai/ALLaM-7B-Instruct-preview"   # modèle de base HF
LORA_DIR     = r"C:\Users\infocom\tradiction_dialect\allam7b_tun_msa_output-3000"
TEST_FILE    = os.path.join(LORA_DIR, "test_set_FINAL.json")
OUTPUT_DIR   = LORA_DIR          # résultats écrits au même endroit
SEED         = 42

os.environ["CUDA_VISIBLE_DEVICES"] = "0,1"
os.environ["TOKENIZERS_PARALLELISM"] = "false"

SEP = "=" * 60
random.seed(SEED)
gc.collect()
torch.cuda.empty_cache()

print(SEP)
print("  Évaluation ALLaM-7B LoRA — Traduction Tunisien → MSA")
print(SEP)
print(f"  PyTorch       : {torch.__version__}")
print(f"  CUDA dispo    : {torch.cuda.is_available()}")
print(f"  Nombre GPUs   : {torch.cuda.device_count()}")
if torch.cuda.is_available():
    for i in range(torch.cuda.device_count()):
        name = torch.cuda.get_device_name(i)
        vram = torch.cuda.get_device_properties(i).total_memory / 1e9
        print(f"  GPU {i}         : {name}  ({vram:.1f} Go VRAM)")
print(SEP + "\n")

# ══════════════════════════════════════════════════════════════
# 1. CHARGEMENT DU TEST SET
# ══════════════════════════════════════════════════════════════
if not os.path.exists(TEST_FILE):
    raise FileNotFoundError(
        f"❌ Test set introuvable : {TEST_FILE}\n"
        "   Assurez-vous que test_set_FINAL.json est dans LORA_DIR."
    )

with open(TEST_FILE, "r", encoding="utf-8") as f:
    test_data = json.load(f)

print(f"✓ Test set chargé : {len(test_data)} exemples\n")

# ══════════════════════════════════════════════════════════════
# 2. CHARGEMENT MODÈLE DE BASE + ADAPTEUR LoRA
# ══════════════════════════════════════════════════════════════
print(f"📥 Chargement modèle de base : {MODEL_ID} ...")

bnb_config = BitsAndBytesConfig(
    load_in_4bit=True,
    bnb_4bit_quant_type="nf4",
    bnb_4bit_compute_dtype=torch.bfloat16,
    bnb_4bit_use_double_quant=True,
)

base_model = AutoModelForCausalLM.from_pretrained(
    MODEL_ID,
    quantization_config=bnb_config,
    device_map="auto",
    trust_remote_code=True,
    torch_dtype=torch.bfloat16,
)
base_model.config.use_cache = True   # activer le KV-cache en inférence

print(f"📥 Chargement adapteur LoRA : {LORA_DIR} ...")
model = PeftModel.from_pretrained(base_model, LORA_DIR)
model.eval()

# ── FIX DTYPE ────────────────────────────────────────────────
# Certains modules (lm_head, embed_tokens) restent en float32
# après chargement PEFT ; on les recaste explicitement en bfloat16.
for name, module in model.named_modules():
    if hasattr(module, "weight") and module.weight is not None:
        if module.weight.dtype == torch.float32:
            module.weight.data = module.weight.data.to(torch.bfloat16)
    if hasattr(module, "bias") and module.bias is not None:
        if module.bias.dtype == torch.float32:
            module.bias.data = module.bias.data.to(torch.bfloat16)

print("✓ Modèle + LoRA chargés (dtype forcé bfloat16)\n")

tokenizer = AutoTokenizer.from_pretrained(
    LORA_DIR,                 # le tokenizer a été sauvegardé avec le modèle
    trust_remote_code=True,
    padding_side="right",
)
if tokenizer.pad_token is None:
    tokenizer.pad_token = tokenizer.eos_token

# ══════════════════════════════════════════════════════════════
# 3. TEMPLATE D'INFÉRENCE (identique au script d'entraînement)
# ══════════════════════════════════════════════════════════════
INFERENCE_TEMPLATE = (
    "### Instruction:\n{instruction}\n\n"
    "### Input:\n{input}\n\n"
    "### Response:\n"
)

# ══════════════════════════════════════════════════════════════
# 4. FONCTIONS D'INFÉRENCE ET DE MÉTRIQUES
# ══════════════════════════════════════════════════════════════

def translate_to_msa(item: dict, max_new_tokens: int = 128) -> str:
    prompt = INFERENCE_TEMPLATE.format(
        instruction=item["instruction"].strip(),
        input=item["input"].strip(),
    )
    # tokenisation + cast bfloat16 sur le device du modèle
    inputs = tokenizer(prompt, return_tensors="pt")
    inputs = {k: v.to(model.device) for k, v in inputs.items()}

    with torch.no_grad():
        # cast input_ids en long (sécurité), attention_mask en long
        outputs = model.generate(
            input_ids=inputs["input_ids"].long(),
            attention_mask=inputs["attention_mask"].long(),
            max_new_tokens=max_new_tokens,
            do_sample=False,
            repetition_penalty=1.1,
            pad_token_id=tokenizer.eos_token_id,
            eos_token_id=tokenizer.eos_token_id,
        )
    new_tokens = outputs[0][inputs["input_ids"].shape[1]:]
    return tokenizer.decode(new_tokens, skip_special_tokens=True).strip()


def is_exact_match(pred: str, ref: str) -> bool:
    return pred.strip().rstrip("،.") == ref.strip().rstrip("،.")


_bleu_metric = BLEU(effective_order=True, tokenize="char")
_ter_metric  = TER()
_chrf_metric = CHRF(word_order=2)
_rouge       = rouge_scorer.RougeScorer(
    ["rouge1", "rouge2", "rougeL"], use_stemmer=False
)


def compute_sentence_bleu(pred, ref):
    try:    return _bleu_metric.sentence_score(pred, [ref]).score
    except: return 0.0

def compute_ter(pred, ref):
    try:    return _ter_metric.sentence_score(pred, [ref]).score
    except: return 100.0

def compute_chrf(pred, ref):
    try:    return _chrf_metric.sentence_score(pred, [ref]).score
    except: return 0.0

def compute_rouge(pred, ref):
    try:
        s = _rouge.score(ref, pred)
        return {"rouge1": s["rouge1"].fmeasure,
                "rouge2": s["rouge2"].fmeasure,
                "rougeL": s["rougeL"].fmeasure}
    except:
        return {"rouge1": 0.0, "rouge2": 0.0, "rougeL": 0.0}

def compute_token_f1(pred, ref):
    pt, rt = set(pred.strip().split()), set(ref.strip().split())
    if not pt or not rt:
        return float(pt == rt)
    common = pt & rt
    if not common:
        return 0.0
    p = len(common) / len(pt)
    r = len(common) / len(rt)
    return 2 * p * r / (p + r)


# ══════════════════════════════════════════════════════════════
# 5. BOUCLE D'ÉVALUATION
# ══════════════════════════════════════════════════════════════
print(SEP)
print("  📊 ÉVALUATION SUR LE TEST SET COMPLET")
print(SEP)

results_eval = []
all_preds, all_refs = [], []
N = len(test_data)

for i, item in enumerate(test_data, 1):
    pred  = translate_to_msa(item)
    ref   = item["output"].strip()
    match = is_exact_match(pred, ref)

    bleu_s  = compute_sentence_bleu(pred, ref)
    ter_s   = compute_ter(pred, ref)
    chrf_s  = compute_chrf(pred, ref)
    rouge_s = compute_rouge(pred, ref)
    tf1_s   = compute_token_f1(pred, ref)

    all_preds.append(pred)
    all_refs.append(ref)

    results_eval.append({
        "id"             : i,
        "instruction"    : item["instruction"],
        "input_tun"      : item["input"],
        "reference"      : ref,
        "predicted"      : pred,
        "exact_match"    : match,
        "token_f1"       : round(tf1_s, 4),
        "bleu"           : round(bleu_s, 2),
        "chrf_pp"        : round(chrf_s, 2),
        "ter"            : round(ter_s, 2),
        "rouge1"         : round(rouge_s["rouge1"], 4),
        "rouge2"         : round(rouge_s["rouge2"], 4),
        "rougeL"         : round(rouge_s["rougeL"], 4),
        "expert_score"   : "",
        "expert_comment" : "",
    })

    symbol = "✓" if match else "✗"
    if i <= 10 or i % 50 == 0:
        print(f"\nQ{i:4d} {symbol}  BLEU={bleu_s:.1f}  chrF++={chrf_s:.1f}"
              f"  TER={ter_s:.1f}  F1={tf1_s:.3f}")
        print(f"  Tunisien  : {item['input']}")
        print(f"  Référence : {ref}")
        print(f"  Prédit    : {pred}")

# ── Métriques corpus ─────────────────────────────────────────
corpus_bleu = BLEU(tokenize="char").corpus_score(all_preds, [all_refs])
corpus_chrf = CHRF(word_order=2).corpus_score(all_preds, [all_refs])
corpus_ter  = TER().corpus_score(all_preds, [all_refs])

exact_matches = sum(r["exact_match"] for r in results_eval)
exact_acc     = exact_matches / N
avg_token_f1  = np.mean([r["token_f1"] for r in results_eval])
avg_rouge1    = np.mean([r["rouge1"]   for r in results_eval])
avg_rouge2    = np.mean([r["rouge2"]   for r in results_eval])
avg_rougeL    = np.mean([r["rougeL"]  for r in results_eval])

summary_metrics = {
    "model"           : MODEL_ID,
    "lora_adapter"    : LORA_DIR,
    "task"            : "Traduction dialecte tunisien → MSA",
    "n_test"          : N,
    "exact_match_acc" : round(float(exact_acc), 4),
    "avg_token_f1"    : round(float(avg_token_f1), 4),
    "corpus_bleu"     : round(corpus_bleu.score, 2),
    "corpus_chrf_pp"  : round(corpus_chrf.score, 2),
    "corpus_ter"      : round(corpus_ter.score, 2),
    "avg_rouge1"      : round(float(avg_rouge1), 4),
    "avg_rouge2"      : round(float(avg_rouge2), 4),
    "avg_rougeL"      : round(float(avg_rougeL), 4),
}

print("\n" + SEP)
print("  🎯 RÉSUMÉ DES MÉTRIQUES (corpus)")
print(SEP)
print(f"  Exact Match Accuracy : {exact_acc*100:.2f}%  ({exact_matches}/{N})")
print(f"  Token F1 (avg)       : {avg_token_f1:.4f}")
print(f"  BLEU (corpus)        : {corpus_bleu.score:.2f}")
print(f"  chrF++ (corpus)      : {corpus_chrf.score:.2f}")
print(f"  TER (corpus)         : {corpus_ter.score:.2f}  (↓ meilleur)")
print(f"  ROUGE-1 (avg)        : {avg_rouge1:.4f}")
print(f"  ROUGE-2 (avg)        : {avg_rouge2:.4f}")
print(f"  ROUGE-L (avg)        : {avg_rougeL:.4f}")
print(SEP)

# ══════════════════════════════════════════════════════════════
# 6. SAUVEGARDE JSON
# ══════════════════════════════════════════════════════════════
eval_summary = {**summary_metrics, "predictions": results_eval}
results_path = os.path.join(OUTPUT_DIR, "evaluation_results.json")
with open(results_path, "w", encoding="utf-8") as f:
    json.dump(eval_summary, f, ensure_ascii=False, indent=2)
print(f"\n💾 Résultats JSON → {results_path}")

# ══════════════════════════════════════════════════════════════
# 7. EXPORT EXCEL MULTI-ONGLETS
# ══════════════════════════════════════════════════════════════

def _header_style(cell, bg="1F4E79"):
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    thin = Side(style="thin", color="FFFFFF")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def _data_style(cell, wrap=False):
    cell.font      = Font(name="Arial", size=9)
    cell.alignment = Alignment(vertical="center", wrap_text=wrap,
                                horizontal="right")
    thin = Side(style="thin", color="D9D9D9")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def _pct_bar(ws, col_letter, row_start, row_end):
    rule = DataBarRule(start_type="num", start_value=0,
                       end_type="num",   end_value=1,
                       color="4472C4")
    ws.conditional_formatting.add(
        f"{col_letter}{row_start}:{col_letter}{row_end}", rule
    )


def export_excel(results, summary, out_path):
    wb = Workbook()

    # ── ONGLET 1 : Métriques globales ────────────────────────
    ws_sum = wb.active
    ws_sum.title = "📋 Métriques Globales"
    ws_sum.sheet_view.rightToLeft = False

    ws_sum.merge_cells("A1:C1")
    title_cell = ws_sum["A1"]
    title_cell.value = "ALLaM-7B LoRA — Évaluation Traduction Tunisien → MSA"
    title_cell.font  = Font(bold=True, size=14, color="FFFFFF", name="Arial")
    title_cell.fill  = PatternFill("solid", start_color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 30

    kv_rows = [
        ("Modèle de base",       summary["model"]),
        ("Adapteur LoRA",        summary["lora_adapter"]),
        ("Tâche",                summary["task"]),
        ("Nombre exemples test", summary["n_test"]),
        ("Date évaluation",
         datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Exact Match Accuracy",
         f"{summary['exact_match_acc']*100:.2f} %"),
        ("Token F1 (avg)",       f"{summary['avg_token_f1']:.4f}"),
        ("BLEU (corpus)",        f"{summary['corpus_bleu']:.2f}"),
        ("chrF++ (corpus)",      f"{summary['corpus_chrf_pp']:.2f}"),
        ("TER (corpus, ↓)",      f"{summary['corpus_ter']:.2f}"),
        ("ROUGE-1 (avg)",        f"{summary['avg_rouge1']:.4f}"),
        ("ROUGE-2 (avg)",        f"{summary['avg_rouge2']:.4f}"),
        ("ROUGE-L (avg)",        f"{summary['avg_rougeL']:.4f}"),
    ]
    for row_idx, (k, v) in enumerate(kv_rows, 2):
        cell_k = ws_sum.cell(row=row_idx, column=1, value=k)
        cell_v = ws_sum.cell(row=row_idx, column=2, value=v)
        if k:
            cell_k.font = Font(bold=True, name="Arial", size=10)
            cell_v.font = Font(name="Arial", size=10)

    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 50

    # ── ONGLET 2 : Évaluation expert ─────────────────────────
    ws_exp = wb.create_sheet("🔍 Évaluation Expert")
    ws_exp.sheet_view.rightToLeft = False
    ws_exp.freeze_panes = "A2"

    expert_headers = [
        "ID", "Texte Tunisien (input)",
        "Référence MSA", "Traduction ALLaM",
        "Exact Match",
        "Token F1", "BLEU", "chrF++", "TER",
        "ROUGE-1", "ROUGE-2", "ROUGE-L",
        "Score Expert (0–5)", "Commentaire Expert",
    ]
    for col, h in enumerate(expert_headers, 1):
        cell = ws_exp.cell(row=1, column=col, value=h)
        bg   = "1F4E79" if col <= 4 else ("2E75B6" if col <= 12 else "C55A11")
        _header_style(cell, bg=bg)

    score_guide = (
        "Guide de notation expert :\n"
        "5 — Traduction parfaite, équivalente au natif MSA\n"
        "4 — Très bonne, quelques imprécisions mineures\n"
        "3 — Acceptable, sens correct mais style perfectible\n"
        "2 — Partielle, perte d'information notable\n"
        "1 — Mauvaise, sens très altéré\n"
        "0 — Complètement erronée ou hors sujet"
    )
    ws_exp["M1"].comment = Comment(score_guide, "ALLaM Evaluation")

    for r_idx, row in enumerate(results, 2):
        match_val  = "✓" if row["exact_match"] else "✗"
        fill_color = "E2EFDA" if row["exact_match"] else "FCE4D6"
        data_row = [
            row["id"], row["input_tun"], row["reference"], row["predicted"],
            match_val,
            row["token_f1"], row["bleu"], row["chrf_pp"], row["ter"],
            row["rouge1"], row["rouge2"], row["rougeL"],
            row["expert_score"], row["expert_comment"],
        ]
        for c_idx, val in enumerate(data_row, 1):
            cell = ws_exp.cell(row=r_idx, column=c_idx, value=val)
            _data_style(cell, wrap=(c_idx in (2, 3, 4, 14)))
            if c_idx in (2, 3, 4, 5):
                cell.fill = PatternFill("solid", start_color=fill_color)

    n_rows = len(results) + 1
    _pct_bar(ws_exp, "F", 2, n_rows)
    _pct_bar(ws_exp, "J", 2, n_rows)
    _pct_bar(ws_exp, "K", 2, n_rows)
    _pct_bar(ws_exp, "L", 2, n_rows)

    dv = DataValidation(type="whole", operator="between",
                        formula1=0, formula2=5,
                        showErrorMessage=True,
                        errorTitle="Valeur invalide",
                        error="Entrez un entier entre 0 et 5.")
    dv.sqref = f"M2:M{n_rows}"
    ws_exp.add_data_validation(dv)

    col_widths = {
        "A": 6,  "B": 40, "C": 40, "D": 40,
        "E": 12, "F": 11, "G": 9,  "H": 10, "I": 9,
        "J": 10, "K": 10, "L": 10, "M": 18, "N": 40,
    }
    for col, w in col_widths.items():
        ws_exp.column_dimensions[col].width = w
    for r in range(2, n_rows + 1):
        ws_exp.row_dimensions[r].height = 60

    # ── ONGLET 3 : Échantillon 50 ────────────────────────────
    ws_sample = wb.create_sheet("🎲 Échantillon 50")
    ws_sample.freeze_panes = "A2"

    sample_indices = random.sample(range(len(results)), min(50, len(results)))
    sample_rows    = [results[i] for i in sorted(sample_indices)]

    sample_headers = ["ID", "Texte Tunisien", "Référence MSA",
                      "Traduction ALLaM", "Token F1", "BLEU",
                      "Score Expert (0–5)", "Commentaire Expert"]
    for col, h in enumerate(sample_headers, 1):
        _header_style(ws_sample.cell(row=1, column=col, value=h))

    for r_idx, row in enumerate(sample_rows, 2):
        vals = [
            row["id"], row["input_tun"], row["reference"], row["predicted"],
            row["token_f1"], row["bleu"],
            row["expert_score"], row["expert_comment"],
        ]
        fill_color = "E2EFDA" if row["exact_match"] else "FFFFFF"
        for c_idx, val in enumerate(vals, 1):
            cell = ws_sample.cell(row=r_idx, column=c_idx, value=val)
            _data_style(cell, wrap=(c_idx in (2, 3, 4, 8)))
            if c_idx in (2, 3, 4):
                cell.fill = PatternFill("solid", start_color=fill_color)

    sample_widths = {"A": 6, "B": 42, "C": 42, "D": 42,
                     "E": 11, "F": 9, "G": 18, "H": 40}
    for col, w in sample_widths.items():
        ws_sample.column_dimensions[col].width = w
    for r in range(2, len(sample_rows) + 2):
        ws_sample.row_dimensions[r].height = 60

    dv2 = DataValidation(type="whole", operator="between",
                         formula1=0, formula2=5,
                         showErrorMessage=True,
                         errorTitle="Valeur invalide",
                         error="Entrez un entier entre 0 et 5.")
    dv2.sqref = f"G2:G{len(sample_rows)+1}"
    ws_sample.add_data_validation(dv2)

    # ── ONGLET 4 : Distribution métriques ────────────────────
    ws_dist = wb.create_sheet("📈 Distribution")

    def _bucket_metric(vals, metric_name, buckets, row_start):
        ws_dist.cell(row_start, 1, metric_name).font = Font(
            bold=True, name="Arial", size=11)
        headers = ["Intervalle", "Nombre", "% du total"]
        for c, h in enumerate(headers, 1):
            _header_style(ws_dist.cell(row_start + 1, c))
            ws_dist.cell(row_start + 1, c, h)
        counts = [sum(1 for v in vals if lo <= v < hi) for lo, hi in buckets]
        for r, ((lo, hi), cnt) in enumerate(zip(buckets, counts),
                                             row_start + 2):
            ws_dist.cell(r, 1, f"[{lo:.2f}, {hi:.2f})")
            ws_dist.cell(r, 2, cnt)
            ws_dist.cell(r, 3,
                f"={get_column_letter(2)}{r}/"
                f"SUM({get_column_letter(2)}{row_start+2}:"
                f"{get_column_letter(2)}{r+len(buckets)-1})")
            ws_dist.cell(r, 3).number_format = "0.0%"

    buckets_01 = [(i / 10, (i + 1) / 10) for i in range(10)]
    _bucket_metric([r["token_f1"]        for r in results],
                   "Token F1",       buckets_01, 1)
    _bucket_metric([r["bleu"] / 100      for r in results],
                   "BLEU (norm.)",   buckets_01, 15)
    _bucket_metric([r["chrf_pp"] / 100   for r in results],
                   "chrF++",         buckets_01, 29)
    _bucket_metric([r["rouge1"]          for r in results],
                   "ROUGE-1",        buckets_01, 43)

    ws_dist.column_dimensions["A"].width = 18
    ws_dist.column_dimensions["B"].width = 12
    ws_dist.column_dimensions["C"].width = 12

    wb.save(out_path)
    print(f"📊 Fichier Excel généré → {out_path}")


# ── Appel export ──────────────────────────────────────────────
excel_path = os.path.join(OUTPUT_DIR, "allam_evaluation_expert.xlsx")
export_excel(results_eval, summary_metrics, excel_path)

print(f"\n💾 JSON  → {results_path}")
print(f"📊 Excel → {excel_path}")
print(f"📄 Test  → {TEST_FILE}")
print("\n" + SEP)
print("  ✅ ÉVALUATION ALLaM-7B TERMINÉE AVEC SUCCÈS!")
print(SEP)
