"""
evaluate_arabart.py
--------------------
Évaluation du modèle AraBART fine-tuné pour la traduction
Dialecte Tunisien → Arabe classique (MSA).

Compatible avec le pipeline finetune_arabart_tun_msa.py :
  - Dataset JSON au format  {"instruction": ..., "input": ..., "output": ...}
  - Métriques : BLEU, chrF++, TER, ROUGE-1/2/L, Token-F1, Exact Match
  - Export JSON + Excel avec onglets Métriques / Expert / Échantillon / Distribution

Prérequis :
    pip install transformers torch sacrebleu rouge-score openpyxl tqdm
"""

import os
import gc
import json
import random
import datetime
import torch
import numpy as np
from tqdm import tqdm
from transformers import AutoTokenizer, AutoModelForSeq2SeqLM
from sacrebleu.metrics import BLEU, TER, CHRF
from rouge_score import rouge_scorer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation


# ── Configuration ──────────────────────────────────────────────────────────────

OUTPUT_DIR   = r"C:\Users\infocom\tradiction_dialect\arabart_tun_msa_output-3000"
DATASET_PATH = r"C:\Users\infocom\tradiction_dialect\dataset_final_3000.json"

RESULTS_JSON  = os.path.join(OUTPUT_DIR, "evaluation_results.json")
RESULTS_EXCEL = os.path.join(OUTPUT_DIR, "arabart_evaluation_expert.xlsx")

MODEL_ID  = "moussaKam/AraBART"   # identique au fine-tuning
MAX_SRC   = 128
MAX_TGT   = 128
NUM_BEAMS = 4
SEED      = 42
SEP       = "=" * 60

DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

random.seed(SEED)


# ── Chargement modèle ──────────────────────────────────────────────────────────

def load_model(model_dir: str):
    """Charge le tokenizer et le modèle fine-tuné depuis model_dir."""
    print(f"📥 Chargement modèle depuis : {model_dir}")
    tokenizer = AutoTokenizer.from_pretrained(model_dir)
    model = AutoModelForSeq2SeqLM.from_pretrained(
        model_dir,
        torch_dtype=torch.bfloat16,
    ).to(DEVICE)
    model.eval()
    gc.collect()
    torch.cuda.empty_cache()
    print(f"  ✓ Modèle chargé sur {DEVICE}\n")
    return tokenizer, model


# ── Chargement dataset ─────────────────────────────────────────────────────────

def load_dataset(path: str) -> list[dict]:
    """
    Lit le fichier JSON du dataset.
    Format attendu : liste de dicts {"instruction": ..., "input": ..., "output": ...}
    Accepte aussi {"data": [...]} comme racine.
    """
    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    data = raw.get("data", raw) if isinstance(raw, dict) else raw
    data = [
        item for item in data
        if isinstance(item, dict)
        and item.get("input", "").strip()
        and item.get("output", "").strip()
    ]
    print(f"✓ Dataset : {len(data)} exemples valides chargés.\n")
    return data


# ── Inférence ──────────────────────────────────────────────────────────────────

def translate_to_msa(item: dict, tokenizer, model) -> str:
    """
    Traduction AraBART avec beam search.
    BART ne nécessite pas de préfixe de tâche : texte brut en entrée.
    """
    src = item["input"].strip()
    inputs = tokenizer(
        src,
        return_tensors="pt",
        max_length=MAX_SRC,
        truncation=True,
    ).to(DEVICE)

    gen_kwargs = dict(
        max_new_tokens=MAX_TGT,
        num_beams=NUM_BEAMS,
        early_stopping=True,
        no_repeat_ngram_size=3,
        length_penalty=1.0,
    )
    if model.config.forced_bos_token_id is not None:
        gen_kwargs["forced_bos_token_id"] = model.config.forced_bos_token_id

    with torch.no_grad():
        outputs = model.generate(**inputs, **gen_kwargs)

    return tokenizer.decode(outputs[0], skip_special_tokens=True).strip()


# ── Métriques phrase ───────────────────────────────────────────────────────────

_bleu_sent = BLEU(effective_order=True, tokenize="char")
_ter_sent  = TER()
_chrf_sent = CHRF(word_order=2)
_rouge     = rouge_scorer.RougeScorer(["rouge1", "rouge2", "rougeL"], use_stemmer=False)


def compute_sentence_bleu(pred: str, ref: str) -> float:
    try:    return _bleu_sent.sentence_score(pred, [ref]).score
    except: return 0.0

def compute_ter(pred: str, ref: str) -> float:
    try:    return _ter_sent.sentence_score(pred, [ref]).score
    except: return 100.0

def compute_chrf(pred: str, ref: str) -> float:
    try:    return _chrf_sent.sentence_score(pred, [ref]).score
    except: return 0.0

def compute_rouge(pred: str, ref: str) -> dict:
    try:
        s = _rouge.score(ref, pred)
        return {k: s[k].fmeasure for k in ("rouge1", "rouge2", "rougeL")}
    except:
        return {"rouge1": 0.0, "rouge2": 0.0, "rougeL": 0.0}

def compute_token_f1(pred: str, ref: str) -> float:
    pt, rt = set(pred.strip().split()), set(ref.strip().split())
    if not pt or not rt:
        return float(pt == rt)
    common = pt & rt
    if not common:
        return 0.0
    prec = len(common) / len(pt)
    rec  = len(common) / len(rt)
    return 2 * prec * rec / (prec + rec)

def is_exact_match(pred: str, ref: str) -> bool:
    return pred.strip().rstrip("،.") == ref.strip().rstrip("،.")


# ── Évaluation complète ────────────────────────────────────────────────────────

def evaluate_dataset(
    test_data: list[dict],
    tokenizer,
    model,
) -> tuple[list[dict], list[str], list[str]]:
    """
    Parcourt test_data, génère les traductions et calcule toutes les métriques.

    Returns:
        (results_eval, all_preds, all_refs)
    """
    print(SEP)
    print("  📊 ÉVALUATION SUR LE TEST SET COMPLET")
    print(SEP)

    results_eval = []
    all_preds, all_refs = [], []
    N = len(test_data)

    for i, item in enumerate(tqdm(test_data, desc="Évaluation"), 1):
        pred  = translate_to_msa(item, tokenizer, model)
        ref   = item["output"].strip()
        match = is_exact_match(pred, ref)

        bleu_s  = compute_sentence_bleu(pred, ref)
        ter_s   = compute_ter(pred, ref)
        chrf_s  = compute_chrf(pred, ref)
        rouge_s = compute_rouge(pred, ref)
        tf1_s   = compute_token_f1(pred, ref)

        all_preds.append(pred)
        all_refs.append(ref)

        results_eval.append({
            "id":             i,
            "instruction":    item.get("instruction", ""),
            "input_tun":      item["input"],
            "reference":      ref,
            "predicted":      pred,
            "exact_match":    match,
            "token_f1":       round(tf1_s, 4),
            "bleu":           round(bleu_s, 2),
            "chrf_pp":        round(chrf_s, 2),
            "ter":            round(ter_s, 2),
            "rouge1":         round(rouge_s["rouge1"], 4),
            "rouge2":         round(rouge_s["rouge2"], 4),
            "rougeL":         round(rouge_s["rougeL"], 4),
            "expert_score":   "",
            "expert_comment": "",
        })

        if i <= 10 or i % 100 == 0:
            sym = "✓" if match else "✗"
            print(f"\nQ{i:4d} {sym}  BLEU={bleu_s:.1f}  chrF++={chrf_s:.1f}"
                  f"  TER={ter_s:.1f}  F1={tf1_s:.3f}")
            print(f"  Tunisien  : {item['input']}")
            print(f"  Référence : {ref}")
            print(f"  Prédit    : {pred}")

    return results_eval, all_preds, all_refs


# ── Métriques corpus ───────────────────────────────────────────────────────────

def compute_corpus_metrics(
    results_eval: list[dict],
    all_preds: list[str],
    all_refs: list[str],
) -> dict:
    """Calcule et affiche les métriques agrégées sur tout le corpus."""
    N = len(results_eval)

    corpus_bleu = BLEU(tokenize="char").corpus_score(all_preds, [all_refs])
    corpus_chrf = CHRF(word_order=2).corpus_score(all_preds, [all_refs])
    corpus_ter  = TER().corpus_score(all_preds, [all_refs])

    exact_matches = sum(r["exact_match"] for r in results_eval)
    exact_acc     = exact_matches / N
    avg_tf1       = np.mean([r["token_f1"] for r in results_eval])
    avg_rouge1    = np.mean([r["rouge1"]   for r in results_eval])
    avg_rouge2    = np.mean([r["rouge2"]   for r in results_eval])
    avg_rougeL    = np.mean([r["rougeL"]   for r in results_eval])

    summary = {
        "model":           MODEL_ID,
        "task":            "Traduction dialecte tunisien → MSA",
        "n_test":          N,
        "exact_match_acc": round(float(exact_acc), 4),
        "avg_token_f1":    round(float(avg_tf1), 4),
        "corpus_bleu":     round(corpus_bleu.score, 2),
        "corpus_chrf_pp":  round(corpus_chrf.score, 2),
        "corpus_ter":      round(corpus_ter.score, 2),
        "avg_rouge1":      round(float(avg_rouge1), 4),
        "avg_rouge2":      round(float(avg_rouge2), 4),
        "avg_rougeL":      round(float(avg_rougeL), 4),
    }

    print("\n" + SEP)
    print("  🎯 RÉSUMÉ DES MÉTRIQUES — AraBART (corpus)")
    print(SEP)
    print(f"  Exact Match Accuracy : {exact_acc*100:.2f}%  ({exact_matches}/{N})")
    print(f"  Token F1 (avg)       : {avg_tf1:.4f}")
    print(f"  BLEU (corpus)        : {corpus_bleu.score:.2f}")
    print(f"  chrF++ (corpus)      : {corpus_chrf.score:.2f}")
    print(f"  TER (corpus)         : {corpus_ter.score:.2f}  (↓ meilleur)")
    print(f"  ROUGE-1 (avg)        : {avg_rouge1:.4f}")
    print(f"  ROUGE-2 (avg)        : {avg_rouge2:.4f}")
    print(f"  ROUGE-L (avg)        : {avg_rougeL:.4f}")
    print(SEP)

    return summary


# ── Export JSON ────────────────────────────────────────────────────────────────

def save_json(results_eval: list[dict], summary: dict, path: str) -> None:
    payload = {**summary, "predictions": results_eval}
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"\n💾 JSON → {path}")


# ── Export Excel ───────────────────────────────────────────────────────────────

def _header_style(cell, bg: str = "1F4E79") -> None:
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFFFFF")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def _data_style(cell, wrap: bool = False) -> None:
    cell.font      = Font(name="Arial", size=9)
    cell.alignment = Alignment(vertical="center", wrap_text=wrap, horizontal="right")
    thin = Side(style="thin", color="D9D9D9")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def _pct_bar(ws, col_letter: str, r1: int, r2: int) -> None:
    ws.conditional_formatting.add(
        f"{col_letter}{r1}:{col_letter}{r2}",
        DataBarRule(start_type="num", start_value=0,
                    end_type="num", end_value=1, color="4472C4"),
    )

def export_excel(results: list[dict], summary: dict, out_path: str) -> None:
    """Génère un fichier Excel avec 4 onglets : Métriques / Expert / Échantillon / Distribution."""
    wb  = Workbook()
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

    # ── Onglet 1 : Métriques globales ────────────────────────────
    ws = wb.active
    ws.title = "📋 Métriques Globales"
    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = "AraBART — Évaluation Traduction Tunisien → MSA"
    c.font  = Font(bold=True, size=14, color="FFFFFF", name="Arial")
    c.fill  = PatternFill("solid", start_color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    rows = [
        ("Modèle",               summary["model"]),
        ("Tâche",                summary["task"]),
        ("Exemples test",        summary["n_test"]),
        ("Date",                 now),
        ("", ""),
        ("Exact Match Accuracy", f"{summary['exact_match_acc']*100:.2f} %"),
        ("Token F1 (avg)",       f"{summary['avg_token_f1']:.4f}"),
        ("BLEU (corpus)",        f"{summary['corpus_bleu']:.2f}"),
        ("chrF++ (corpus)",      f"{summary['corpus_chrf_pp']:.2f}"),
        ("TER (corpus, ↓)",      f"{summary['corpus_ter']:.2f}"),
        ("ROUGE-1 (avg)",        f"{summary['avg_rouge1']:.4f}"),
        ("ROUGE-2 (avg)",        f"{summary['avg_rouge2']:.4f}"),
        ("ROUGE-L (avg)",        f"{summary['avg_rougeL']:.4f}"),
    ]
    for ri, (k, v) in enumerate(rows, 2):
        ck = ws.cell(ri, 1, k)
        cv = ws.cell(ri, 2, v)
        if k:
            ck.font = Font(bold=True, name="Arial", size=10)
            cv.font = Font(name="Arial", size=10)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    # ── Onglet 2 : Évaluation expert ─────────────────────────────
    we = wb.create_sheet("🔍 Évaluation Expert")
    we.freeze_panes = "A2"
    headers = [
        "ID", "Texte Tunisien", "Référence MSA", "Traduction AraBART",
        "Exact Match", "Token F1", "BLEU", "chrF++", "TER",
        "ROUGE-1", "ROUGE-2", "ROUGE-L", "Score Expert (0–5)", "Commentaire",
    ]
    for col, h in enumerate(headers, 1):
        bg = "1F4E79" if col <= 4 else ("2E75B6" if col <= 12 else "C55A11")
        _header_style(we.cell(1, col, h), bg=bg)

    we["M1"].comment = Comment(
        "5=parfait · 4=très bon · 3=acceptable · 2=partiel · 1=mauvais · 0=erroné",
        "AraBART Eval",
    )

    for ri, row in enumerate(results, 2):
        fc = "E2EFDA" if row["exact_match"] else "FCE4D6"
        vals = [
            row["id"], row["input_tun"], row["reference"], row["predicted"],
            "✓" if row["exact_match"] else "✗",
            row["token_f1"], row["bleu"], row["chrf_pp"], row["ter"],
            row["rouge1"], row["rouge2"], row["rougeL"],
            row["expert_score"], row["expert_comment"],
        ]
        for ci, val in enumerate(vals, 1):
            cell = we.cell(ri, ci, val)
            _data_style(cell, wrap=(ci in (2, 3, 4, 14)))
            if ci in (2, 3, 4, 5):
                cell.fill = PatternFill("solid", start_color=fc)

    nr = len(results) + 1
    for col in ("F", "J", "K", "L"):
        _pct_bar(we, col, 2, nr)

    dv = DataValidation(
        type="whole", operator="between", formula1=0, formula2=5,
        showErrorMessage=True, errorTitle="Invalide", error="Entrez 0–5.",
    )
    dv.sqref = f"M2:M{nr}"
    we.add_data_validation(dv)

    for col, w in {
        "A": 6, "B": 40, "C": 40, "D": 40, "E": 12, "F": 11,
        "G": 9, "H": 10, "I": 9, "J": 10, "K": 10, "L": 10,
        "M": 18, "N": 40,
    }.items():
        we.column_dimensions[col].width = w
    for r in range(2, nr + 1):
        we.row_dimensions[r].height = 60

    # ── Onglet 3 : Échantillon 50 ────────────────────────────────
    ws3 = wb.create_sheet("🎲 Échantillon 50")
    ws3.freeze_panes = "A2"
    s_idx  = random.sample(range(len(results)), min(50, len(results)))
    s_rows = [results[i] for i in sorted(s_idx)]
    for col, h in enumerate(
        ["ID", "Texte Tunisien", "Référence MSA",
         "Traduction AraBART", "Token F1", "BLEU", "Score Expert", "Commentaire"], 1
    ):
        _header_style(ws3.cell(1, col, h))
    for ri, row in enumerate(s_rows, 2):
        fc = "E2EFDA" if row["exact_match"] else "FFFFFF"
        for ci, val in enumerate(
            [row["id"], row["input_tun"], row["reference"], row["predicted"],
             row["token_f1"], row["bleu"], row["expert_score"], row["expert_comment"]], 1
        ):
            cell = ws3.cell(ri, ci, val)
            _data_style(cell, wrap=(ci in (2, 3, 4, 8)))
            if ci in (2, 3, 4):
                cell.fill = PatternFill("solid", start_color=fc)
    for col, w in {"A": 6, "B": 42, "C": 42, "D": 42,
                   "E": 11, "F": 9, "G": 18, "H": 40}.items():
        ws3.column_dimensions[col].width = w

    # ── Onglet 4 : Distribution ──────────────────────────────────
    wd = wb.create_sheet("📈 Distribution")
    buckets = [(i / 10, (i + 1) / 10) for i in range(10)]

    def _bucket(vals, name, r0):
        wd.cell(r0, 1, name).font = Font(bold=True, name="Arial", size=11)
        for c, h in enumerate(["Intervalle", "Nombre", "% total"], 1):
            _header_style(wd.cell(r0 + 1, c, h))
        counts = [sum(1 for v in vals if lo <= v < hi) for lo, hi in buckets]
        for r, ((lo, hi), cnt) in enumerate(zip(buckets, counts), r0 + 2):
            wd.cell(r, 1, f"[{lo:.2f},{hi:.2f})")
            wd.cell(r, 2, cnt)
            wd.cell(r, 3, f"=B{r}/SUM(B{r0+2}:B{r0+11})")
            wd.cell(r, 3).number_format = "0.0%"

    _bucket([r["token_f1"]     for r in results], "Token F1",    1)
    _bucket([r["bleu"] / 100   for r in results], "BLEU (norm)", 15)
    _bucket([r["chrf_pp"] / 100 for r in results], "chrF++",     29)
    _bucket([r["rouge1"]       for r in results], "ROUGE-1",     43)
    for col, w in {"A": 18, "B": 12, "C": 12}.items():
        wd.column_dimensions[col].width = w

    wb.save(out_path)
    print(f"📊 Excel → {out_path}")


# ── Pipeline principal ─────────────────────────────────────────────────────────

def main():
    print(SEP)
    print("  AraBART — Évaluation Traduction Tunisien → MSA")
    print(SEP)
    print(f"  Device : {DEVICE}")
    print(f"  Modèle : {OUTPUT_DIR}\n")

    # 1. Charger le modèle fine-tuné
    tokenizer, model = load_model(OUTPUT_DIR)

    # 2. Charger le test set (celui sauvegardé pendant le fine-tuning en priorité)
    test_set_path = os.path.join(OUTPUT_DIR, "test_set_FINAL.json")
    if os.path.exists(test_set_path):
        print(f"✓ Utilisation du test set du fine-tuning : {test_set_path}")
        test_data = load_dataset(test_set_path)
    else:
        print("⚠ test_set_FINAL.json introuvable, split depuis DATASET_PATH...")
        all_data  = load_dataset(DATASET_PATH)
        random.seed(SEED)
        random.shuffle(all_data)
        split     = int(len(all_data) * 0.85)
        test_data = all_data[split:]

    # 3. Évaluation phrase par phrase
    results_eval, all_preds, all_refs = evaluate_dataset(test_data, tokenizer, model)

    # 4. Métriques corpus
    summary = compute_corpus_metrics(results_eval, all_preds, all_refs)

    # 5. Exports
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    save_json(results_eval, summary, RESULTS_JSON)
    export_excel(results_eval, summary, RESULTS_EXCEL)

    print("\n" + SEP)
    print("  ✅ ÉVALUATION TERMINÉE")
    print(SEP)


if __name__ == "__main__":
    main()
