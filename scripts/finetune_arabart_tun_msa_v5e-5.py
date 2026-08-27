"""
Fine-tuning AraBART (seq2seq) — Traduction dialecte tunisien → MSA
Modèle : moussaKam/AraBART  (ou UBC-NLP/AraBART)
Méthode : Fine-tuning complet seq2seq (encodeur-décodeur BART)
Dataset : JSON {instruction, input, output}

pip install transformers datasets accelerate torch sacrebleu
            rouge-score openpyxl packaging sentencepiece
"""

import os, gc, json, random, datetime
import torch
import numpy as np
from transformers import (
    AutoTokenizer,
    AutoModelForSeq2SeqLM,
    Seq2SeqTrainer,
    Seq2SeqTrainingArguments,
    DataCollatorForSeq2Seq,
    EarlyStoppingCallback,
)
from datasets import Dataset
from sacrebleu.metrics import BLEU, TER, CHRF
from rouge_score import rouge_scorer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation

# ══════════════════════════════════════════════════════════════
# 0. ENVIRONNEMENT
# ══════════════════════════════════════════════════════════════
os.environ["CUDA_VISIBLE_DEVICES"] = "0,1"
os.environ["TOKENIZERS_PARALLELISM"] = "false"

gc.collect()
torch.cuda.empty_cache()

SEP  = "=" * 60
SEED = 42
random.seed(SEED)

print(SEP)
print("  Fine-tuning AraBART — Traduction Tunisien → MSA")
print(SEP)
print(f"  PyTorch  : {torch.__version__}")
print(f"  CUDA     : {torch.cuda.is_available()}")
print(f"  GPU(s)   : {torch.cuda.device_count()}")
if torch.cuda.is_available():
    for i in range(torch.cuda.device_count()):
        name = torch.cuda.get_device_name(i)
        vram = torch.cuda.get_device_properties(i).total_memory / 1e9
        print(f"  GPU {i}   : {name}  ({vram:.1f} Go)")
print(SEP + "\n")

# ══════════════════════════════════════════════════════════════
# 1. CHEMINS — À ADAPTER
# ══════════════════════════════════════════════════════════════
DATASET_FILE = r"C:\Users\infocom\tradiction_dialect\dataset_final_3000.json"
OUTPUT_DIR   = r"C:\Users\infocom\tradiction_dialect\arabart_tun_msa_output-5e-5"
# Deux options disponibles sur HuggingFace — choisir l'une :
#   "moussaKam/AraBART"   — entraîné spécifiquement sur le dialecte arabe
#   "UBC-NLP/AraBART"     — variante UBC (meilleure base MSA)
MODEL_ID     = "moussaKam/AraBART"

os.makedirs(OUTPUT_DIR, exist_ok=True)

if not os.path.exists(DATASET_FILE):
    raise FileNotFoundError(f"❌ Dataset introuvable : {DATASET_FILE}")
print(f"✓ Dataset trouvé : {DATASET_FILE}\n")

# ══════════════════════════════════════════════════════════════
# 2. CHARGEMENT ET PRÉPARATION DU DATASET
# ══════════════════════════════════════════════════════════════
print("📊 Chargement du dataset...")
with open(DATASET_FILE, "r", encoding="utf-8") as f:
    raw_data = json.load(f)

raw = raw_data.get("data", raw_data) if isinstance(raw_data, dict) else raw_data
raw = [
    item for item in raw
    if isinstance(item, dict)
    and item.get("input", "").strip()
    and item.get("output", "").strip()
]
print(f"  Exemples valides : {len(raw)}")

random.shuffle(raw)
split      = int(len(raw) * 0.85)
train_data = raw[:split]
test_data  = raw[split:]

with open(f"{OUTPUT_DIR}/test_set_FINAL.json", "w", encoding="utf-8") as f:
    json.dump(test_data, f, ensure_ascii=False, indent=2)

print(f"  Train : {len(train_data)} | Test : {len(test_data)}")
print(f"  Test set → {OUTPUT_DIR}/test_set_FINAL.json\n")

# ══════════════════════════════════════════════════════════════
# 3. TOKENIZER + MODÈLE
# AraBART : architecture BART pré-entraînée sur corpus arabe
# Pas de préfixe de tâche nécessaire (contrairement à T5)
# ══════════════════════════════════════════════════════════════
MAX_SRC = 128
MAX_TGT = 128

print(f"📥 Chargement {MODEL_ID}...")
tokenizer = AutoTokenizer.from_pretrained(MODEL_ID)
model     = AutoModelForSeq2SeqLM.from_pretrained(
    MODEL_ID,
    torch_dtype=torch.bfloat16,
)
print(f"  Paramètres : {sum(p.numel() for p in model.parameters()):,}\n")

# ══════════════════════════════════════════════════════════════
# 4. PRÉPARATION DES DONNÉES
# BART : pas de préfixe de tâche, encodeur prend le texte brut
# Le token de forçage du décodeur (forced_bos_token_id) aide
# à orienter la génération vers le MSA si disponible.
# ══════════════════════════════════════════════════════════════

# Forcer le BOS du décodeur si le tokenizer le supporte
if hasattr(tokenizer, "lang_code_to_id"):
    # Certaines versions d'AraBART exposent des codes langue
    forced_bos = tokenizer.lang_code_to_id.get("ar_AR", None)
    if forced_bos:
        model.config.forced_bos_token_id = forced_bos
        print(f"  ✓ forced_bos_token_id = {forced_bos} (ar_AR)\n")

def preprocess(batch):
    """
    Tokenise en mode batched=True.
    batch est un dict de listes : {"input": [...], "output": [...], ...}
    BART : pas de préfixe de tâche, texte brut en entrée.
    """
    sources = [s.strip() for s in batch["input"]]
    targets = [t.strip() for t in batch["output"]]

    model_inputs = tokenizer(
        sources,
        max_length=MAX_SRC,
        truncation=True,
        padding=False,
    )
    # text_target remplace l'API dépréciée as_target_tokenizer
    labels = tokenizer(
        text_target=targets,
        max_length=MAX_TGT,
        truncation=True,
        padding=False,
    )
    model_inputs["labels"] = labels["input_ids"]
    return model_inputs

# Conversion en Dataset HF
print("⚙️  Tokenisation...")
_col_names = list(train_data[0].keys())
train_ds = Dataset.from_list(train_data).map(
    preprocess, batched=True, batch_size=256, remove_columns=_col_names
)
val_ds = Dataset.from_list(test_data[:80]).map(
    preprocess, batched=True, batch_size=256, remove_columns=_col_names
)
print(f"  ✓ Train tokenisé : {len(train_ds)} exemples")
print(f"  ✓ Val tokenisée  : {len(val_ds)} exemples\n")

# ══════════════════════════════════════════════════════════════
# 5. DATA COLLATOR
# ══════════════════════════════════════════════════════════════
data_collator = DataCollatorForSeq2Seq(
    tokenizer,
    model=model,
    label_pad_token_id=-100,
    pad_to_multiple_of=8,
)

# ══════════════════════════════════════════════════════════════
# 6. ARGUMENTS D'ENTRAÎNEMENT
# BART converge généralement plus vite que T5 sur la génération
# ══════════════════════════════════════════════════════════════
training_args = Seq2SeqTrainingArguments(
    output_dir=OUTPUT_DIR,
    per_device_train_batch_size=16,
    per_device_eval_batch_size=16,
    gradient_accumulation_steps=2,
    num_train_epochs=10,
    learning_rate=5e-5,           # LR plus faible que T5 (BART plus sensible)
    lr_scheduler_type="cosine",
    warmup_steps=200,
    weight_decay=0.01,
    bf16=True,
    fp16=False,
    predict_with_generate=True,
    generation_max_length=MAX_TGT,
    logging_steps=20,
    eval_steps=100,
    save_steps=200,
    eval_strategy="steps",
    save_strategy="steps",
    save_total_limit=2,
    load_best_model_at_end=True,
    metric_for_best_model="eval_loss",
    dataloader_num_workers=0,
    report_to="none",
    seed=SEED,
)

# ══════════════════════════════════════════════════════════════
# 7. TRAINER
# ══════════════════════════════════════════════════════════════
trainer = Seq2SeqTrainer(
    model=model,
    args=training_args,
    train_dataset=train_ds,
    eval_dataset=val_ds,
    tokenizer=tokenizer,
    data_collator=data_collator,
    callbacks=[EarlyStoppingCallback(early_stopping_patience=3)],
)

print(SEP)
print("  🔥 DÉBUT ENTRAÎNEMENT AraBART")
print(SEP)
print(f"  Train   : {len(train_data)} exemples")
print(f"  Val     : {len(val_ds)} exemples")
print(f"  Batch   : {16 * 2 * max(1, torch.cuda.device_count())}")
print(f"  Époques : 10 (early stopping patience=3)")
print(f"  LR      : 3e-5 (BART sensible aux LR élevés)")
print(SEP + "\n")

trainer.train()

# ══════════════════════════════════════════════════════════════
# 8. SAUVEGARDE
# ══════════════════════════════════════════════════════════════
print("\n💾 Sauvegarde du modèle...")
trainer.save_model(OUTPUT_DIR)
tokenizer.save_pretrained(OUTPUT_DIR)
print(f"✅ MODÈLE SAUVEGARDÉ → {OUTPUT_DIR}\n")

# ══════════════════════════════════════════════════════════════
# 9. INFÉRENCE
# ══════════════════════════════════════════════════════════════
model.eval()
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
model.to(device)

def translate_to_msa(item: dict, num_beams: int = 4) -> str:
    """Traduction AraBART avec beam search."""
    src = item["input"].strip()
    inputs = tokenizer(
        src,
        return_tensors="pt",
        max_length=MAX_SRC,
        truncation=True,
    ).to(device)

    gen_kwargs = dict(
        max_new_tokens=MAX_TGT,
        num_beams=num_beams,
        early_stopping=True,
        no_repeat_ngram_size=3,
        length_penalty=1.0,
    )
    # Appliquer forced_bos si configuré
    if model.config.forced_bos_token_id is not None:
        gen_kwargs["forced_bos_token_id"] = model.config.forced_bos_token_id

    with torch.no_grad():
        outputs = model.generate(**inputs, **gen_kwargs)
    return tokenizer.decode(outputs[0], skip_special_tokens=True).strip()


# ══════════════════════════════════════════════════════════════
# 10. MÉTRIQUES
# ══════════════════════════════════════════════════════════════
_bleu  = BLEU(effective_order=True, tokenize="char")
_ter   = TER()
_chrf  = CHRF(word_order=2)
_rouge = rouge_scorer.RougeScorer(["rouge1","rouge2","rougeL"], use_stemmer=False)

def compute_sentence_bleu(p, r):
    try:    return _bleu.sentence_score(p, [r]).score
    except: return 0.0

def compute_ter(p, r):
    try:    return _ter.sentence_score(p, [r]).score
    except: return 100.0

def compute_chrf(p, r):
    try:    return _chrf.sentence_score(p, [r]).score
    except: return 0.0

def compute_rouge(p, r):
    try:
        s = _rouge.score(r, p)
        return {k: s[k].fmeasure for k in ("rouge1","rouge2","rougeL")}
    except:
        return {"rouge1":0.0,"rouge2":0.0,"rougeL":0.0}

def compute_token_f1(p, r):
    pt, rt = set(p.strip().split()), set(r.strip().split())
    if not pt or not rt: return float(pt == rt)
    common = pt & rt
    if not common: return 0.0
    prec = len(common)/len(pt); rec = len(common)/len(rt)
    return 2*prec*rec/(prec+rec)

def is_exact_match(p, r):
    return p.strip().rstrip("،.") == r.strip().rstrip("،.")

# ══════════════════════════════════════════════════════════════
# 11. ÉVALUATION
# ══════════════════════════════════════════════════════════════
print(SEP)
print("  📊 ÉVALUATION SUR LE TEST SET COMPLET")
print(SEP)

results_eval = []
all_preds, all_refs = [], []
N = len(test_data)

for i, item in enumerate(test_data, 1):
    pred  = translate_to_msa(item)
    ref   = item["output"].strip()
    match = is_exact_match(pred, ref)

    bleu_s  = compute_sentence_bleu(pred, ref)
    ter_s   = compute_ter(pred, ref)
    chrf_s  = compute_chrf(pred, ref)
    rouge_s = compute_rouge(pred, ref)
    tf1_s   = compute_token_f1(pred, ref)

    all_preds.append(pred)
    all_refs.append(ref)

    results_eval.append({
        "id": i, "instruction": item.get("instruction",""),
        "input_tun": item["input"], "reference": ref, "predicted": pred,
        "exact_match": match,
        "token_f1": round(tf1_s,4), "bleu": round(bleu_s,2),
        "chrf_pp": round(chrf_s,2), "ter": round(ter_s,2),
        "rouge1": round(rouge_s["rouge1"],4),
        "rouge2": round(rouge_s["rouge2"],4),
        "rougeL": round(rouge_s["rougeL"],4),
        "expert_score": "", "expert_comment": "",
    })

    symbol = "✓" if match else "✗"
    if i <= 10 or i % 100 == 0:
        print(f"\nQ{i:4d} {symbol}  BLEU={bleu_s:.1f}  chrF++={chrf_s:.1f}"
              f"  TER={ter_s:.1f}  F1={tf1_s:.3f}")
        print(f"  Tunisien  : {item['input']}")
        print(f"  Référence : {ref}")
        print(f"  Prédit    : {pred}")

# ── Métriques corpus ──────────────────────────────────────────
corpus_bleu = BLEU(tokenize="char").corpus_score(all_preds, [all_refs])
corpus_chrf = CHRF(word_order=2).corpus_score(all_preds, [all_refs])
corpus_ter  = TER().corpus_score(all_preds, [all_refs])

exact_matches = sum(r["exact_match"] for r in results_eval)
exact_acc     = exact_matches / N
avg_tf1    = np.mean([r["token_f1"] for r in results_eval])
avg_rouge1 = np.mean([r["rouge1"]   for r in results_eval])
avg_rouge2 = np.mean([r["rouge2"]   for r in results_eval])
avg_rougeL = np.mean([r["rougeL"]   for r in results_eval])

summary_metrics = {
    "model": MODEL_ID, "task": "Traduction dialecte tunisien → MSA",
    "n_test": N,
    "exact_match_acc": round(float(exact_acc),4),
    "avg_token_f1":    round(float(avg_tf1),4),
    "corpus_bleu":     round(corpus_bleu.score,2),
    "corpus_chrf_pp":  round(corpus_chrf.score,2),
    "corpus_ter":      round(corpus_ter.score,2),
    "avg_rouge1":      round(float(avg_rouge1),4),
    "avg_rouge2":      round(float(avg_rouge2),4),
    "avg_rougeL":      round(float(avg_rougeL),4),
}

print("\n" + SEP)
print("  🎯 RÉSUMÉ DES MÉTRIQUES — AraBART (corpus)")
print(SEP)
print(f"  Exact Match Accuracy : {exact_acc*100:.2f}%  ({exact_matches}/{N})")
print(f"  Token F1 (avg)       : {avg_tf1:.4f}")
print(f"  BLEU (corpus)        : {corpus_bleu.score:.2f}")
print(f"  chrF++ (corpus)      : {corpus_chrf.score:.2f}")
print(f"  TER (corpus)         : {corpus_ter.score:.2f}  (↓ meilleur)")
print(f"  ROUGE-1 (avg)        : {avg_rouge1:.4f}")
print(f"  ROUGE-2 (avg)        : {avg_rouge2:.4f}")
print(f"  ROUGE-L (avg)        : {avg_rougeL:.4f}")
print(SEP)

# ══════════════════════════════════════════════════════════════
# 12. SAUVEGARDE JSON + EXCEL
# ══════════════════════════════════════════════════════════════
eval_summary = {**summary_metrics, "predictions": results_eval}
results_path = f"{OUTPUT_DIR}/evaluation_results.json"
with open(results_path, "w", encoding="utf-8") as f:
    json.dump(eval_summary, f, ensure_ascii=False, indent=2)
print(f"\n💾 Résultats JSON → {results_path}")

# ── Helpers Excel ─────────────────────────────────────────────
def _header_style(cell, bg="1F4E79"):
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFFFFF")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def _data_style(cell, wrap=False):
    cell.font      = Font(name="Arial", size=9)
    cell.alignment = Alignment(vertical="center", wrap_text=wrap, horizontal="right")
    thin = Side(style="thin", color="D9D9D9")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def _pct_bar(ws, col_letter, r1, r2):
    ws.conditional_formatting.add(
        f"{col_letter}{r1}:{col_letter}{r2}",
        DataBarRule(start_type="num", start_value=0,
                    end_type="num", end_value=1, color="4472C4"),
    )

def export_excel(results, summary, out_path):
    wb  = Workbook()
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

    # Onglet 1 — Métriques globales
    ws = wb.active
    ws.title = "📋 Métriques Globales"
    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = "AraBART — Évaluation Traduction Tunisien → MSA"
    c.font  = Font(bold=True, size=14, color="FFFFFF", name="Arial")
    c.fill  = PatternFill("solid", start_color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    rows = [
        ("Modèle",               summary["model"]),
        ("Tâche",                summary["task"]),
        ("Exemples test",        summary["n_test"]),
        ("Date",                 now),
        ("",""),
        ("Exact Match Accuracy", f"{summary['exact_match_acc']*100:.2f} %"),
        ("Token F1 (avg)",       f"{summary['avg_token_f1']:.4f}"),
        ("BLEU (corpus)",        f"{summary['corpus_bleu']:.2f}"),
        ("chrF++ (corpus)",      f"{summary['corpus_chrf_pp']:.2f}"),
        ("TER (corpus, ↓)",      f"{summary['corpus_ter']:.2f}"),
        ("ROUGE-1 (avg)",        f"{summary['avg_rouge1']:.4f}"),
        ("ROUGE-2 (avg)",        f"{summary['avg_rouge2']:.4f}"),
        ("ROUGE-L (avg)",        f"{summary['avg_rougeL']:.4f}"),
    ]
    for ri, (k, v) in enumerate(rows, 2):
        ck = ws.cell(ri, 1, k); cv = ws.cell(ri, 2, v)
        if k:
            ck.font = Font(bold=True, name="Arial", size=10)
            cv.font = Font(name="Arial", size=10)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    # Onglet 2 — Évaluation expert
    we = wb.create_sheet("🔍 Évaluation Expert")
    we.freeze_panes = "A2"
    headers = ["ID","Texte Tunisien","Référence MSA","Traduction AraBART",
               "Exact Match","Token F1","BLEU","chrF++","TER",
               "ROUGE-1","ROUGE-2","ROUGE-L","Score Expert (0–5)","Commentaire"]
    for col, h in enumerate(headers, 1):
        bg = "1F4E79" if col<=4 else ("2E75B6" if col<=12 else "C55A11")
        _header_style(we.cell(1, col, h), bg=bg)

    we["M1"].comment = Comment(
        "5=parfait · 4=très bon · 3=acceptable · 2=partiel · 1=mauvais · 0=erroné",
        "AraBART Eval")

    for ri, row in enumerate(results, 2):
        fc = "E2EFDA" if row["exact_match"] else "FCE4D6"
        vals = [row["id"], row["input_tun"], row["reference"], row["predicted"],
                "✓" if row["exact_match"] else "✗",
                row["token_f1"], row["bleu"], row["chrf_pp"], row["ter"],
                row["rouge1"], row["rouge2"], row["rougeL"],
                row["expert_score"], row["expert_comment"]]
        for ci, val in enumerate(vals, 1):
            cell = we.cell(ri, ci, val)
            _data_style(cell, wrap=(ci in (2,3,4,14)))
            if ci in (2,3,4,5):
                cell.fill = PatternFill("solid", start_color=fc)

    nr = len(results)+1
    for col in ("F","J","K","L"): _pct_bar(we, col, 2, nr)
    dv = DataValidation(type="whole", operator="between", formula1=0, formula2=5,
                        showErrorMessage=True, errorTitle="Invalide",
                        error="Entrez 0–5.")
    dv.sqref = f"M2:M{nr}"; we.add_data_validation(dv)
    for col, w in {"A":6,"B":40,"C":40,"D":40,"E":12,"F":11,"G":9,"H":10,
                   "I":9,"J":10,"K":10,"L":10,"M":18,"N":40}.items():
        we.column_dimensions[col].width = w
    for r in range(2, nr+1): we.row_dimensions[r].height = 60

    # Onglet 3 — Échantillon 50
    ws3 = wb.create_sheet("🎲 Échantillon 50")
    ws3.freeze_panes = "A2"
    s_idx  = random.sample(range(len(results)), min(50, len(results)))
    s_rows = [results[i] for i in sorted(s_idx)]
    for col, h in enumerate(["ID","Texte Tunisien","Référence MSA",
                              "Traduction AraBART","Token F1","BLEU",
                              "Score Expert","Commentaire"], 1):
        _header_style(ws3.cell(1, col, h))
    for ri, row in enumerate(s_rows, 2):
        fc = "E2EFDA" if row["exact_match"] else "FFFFFF"
        for ci, val in enumerate([row["id"],row["input_tun"],row["reference"],
                                   row["predicted"],row["token_f1"],row["bleu"],
                                   row["expert_score"],row["expert_comment"]], 1):
            cell = ws3.cell(ri, ci, val)
            _data_style(cell, wrap=(ci in (2,3,4,8)))
            if ci in (2,3,4): cell.fill = PatternFill("solid", start_color=fc)
    for col, w in {"A":6,"B":42,"C":42,"D":42,"E":11,"F":9,"G":18,"H":40}.items():
        ws3.column_dimensions[col].width = w

    # Onglet 4 — Distribution
    wd = wb.create_sheet("📈 Distribution")
    buckets = [(i/10,(i+1)/10) for i in range(10)]
    def _bucket(vals, name, r0):
        wd.cell(r0,1,name).font = Font(bold=True,name="Arial",size=11)
        for c, h in enumerate(["Intervalle","Nombre","% total"],1):
            _header_style(wd.cell(r0+1,c,h))
        for r,((lo,hi),cnt) in enumerate(
                zip(buckets,[sum(1 for v in vals if lo<=v<hi) for lo,hi in buckets]),
                r0+2):
            wd.cell(r,1,f"[{lo:.2f},{hi:.2f})")
            wd.cell(r,2,cnt)
            wd.cell(r,3,f"=B{r}/SUM(B{r0+2}:B{r0+11})")
            wd.cell(r,3).number_format = "0.0%"
    _bucket([r["token_f1"]    for r in results], "Token F1",    1)
    _bucket([r["bleu"]/100    for r in results], "BLEU (norm)", 15)
    _bucket([r["chrf_pp"]/100 for r in results], "chrF++",      29)
    _bucket([r["rouge1"]      for r in results], "ROUGE-1",     43)
    for col, w in {"A":18,"B":12,"C":12}.items():
        wd.column_dimensions[col].width = w

    wb.save(out_path)
    print(f"📊 Excel → {out_path}")

excel_path = f"{OUTPUT_DIR}/arabart_evaluation_expert.xlsx"
export_excel(results_eval, summary_metrics, excel_path)

with open(results_path, "w", encoding="utf-8") as f:
    json.dump(eval_summary, f, ensure_ascii=False, indent=2)

print(f"\n💾 JSON  → {results_path}")
print(f"📊 Excel → {excel_path}")
print("\n" + SEP)
print("  ✅ FINE-TUNING ET ÉVALUATION AraBART TERMINÉS!")
print(SEP)
